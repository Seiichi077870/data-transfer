# ================================================================================
# 流用元無しピッキングリスト自動生成プログラム（試行錯誤ロジック実装版）
#
# 作成日: 2025/01/15
# 仕様: Excel VBA版を完全移植（2段階パターン割り当て方式）
#
# フォルダ構成:
#   C:\temp\python_流用元無し_Pick\
#   ├── 7552224482_xxx.xlsx                    ← 処理対象ファイル
#   ├── 7552224482_xxx_Pick完了_250115.xlsx    ← 出力ファイル
#   └── main.py                                ← このファイル1
#
#   C:\temp\Newピッキング_対象照会\参照先\      ← 参照マスタ（読取専用）
#   ├── T_VJ_ＣＭピッキングマスタ.xlsx
#   ├── T_VJ_部品出庫先テーブル.xlsx
#   └── New_部品ピッキング諸元.xlsx
# ================================================================================

import sys
from pathlib import Path
from datetime import datetime
import re

# ================================================================================
# 外部ライブラリパス設定
# ================================================================================
# 相対パスで指定（どのPCでも動作）
LIBS_PATH = Path(__file__).parent.parent / "libs"

if LIBS_PATH.exists():
    if str(LIBS_PATH) not in sys.path:
        sys.path.insert(0, str(LIBS_PATH))
    print(f"✅ 外部ライブラリパスを読み込みました: {LIBS_PATH}")
else:
    print(f"❌ 外部ライブラリフォルダが見つかりません: {LIBS_PATH}")
    sys.exit(1)

# ================================================================================
# パッケージインポート
# ================================================================================
try:
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    print("✅ ライブラリインポート成功\n")
except ImportError as e:
    print(f"❌ インポートエラー: {e}")
    sys.exit(1)


# ================================================================================
# ロガークラス
# ================================================================================
class PickingLogger:
    """処理ログを記録するクラス"""

    def __init__(self):
        self.logs = []
        self.warnings = []
        self.part_decisions = []
        self.lane_decisions = []  # 出庫先ごとの判定根拠

    def add_step(self, step_name, details):
        """処理ステップを記録"""
        self.logs.append({
            'タイムスタンプ': datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
            'ステップ': step_name,
            '詳細': str(details)
        })

    def add_warning(self, message, details=""):
        """警告を記録"""
        self.warnings.append({
            'タイムスタンプ': datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
            '警告内容': message,
            '詳細': str(details)
        })

    def add_lane_decision(self, lane, required_parts, selected_pattern, match_count, missing_count, excess_count,
                          reason):
        """出庫先ごとの判定根拠を記録"""
        self.lane_decisions.append({
            '出庫先': lane,
            '必要部品数': len(required_parts),
            '必要部品リスト': ", ".join(sorted(required_parts)),
            '選択パターン': selected_pattern,
            '一致数': match_count,
            '不足数': missing_count,
            '余剰数': excess_count,
            '理由': reason
        })

    def export_to_excel(self, writer):
        """ログをExcelに出力"""
        if self.logs:
            df_logs = pd.DataFrame(self.logs)
            df_logs.to_excel(writer, sheet_name='処理ログ', index=False)

        if self.warnings:
            df_warnings = pd.DataFrame(self.warnings)
            df_warnings.to_excel(writer, sheet_name='警告リスト', index=False)

        if self.lane_decisions:
            df_decisions = pd.DataFrame(self.lane_decisions)
            df_decisions.to_excel(writer, sheet_name='出庫先別判定根拠', index=False)

# ================================================================================
# メインクラス
# ================================================================================
class PickingListGenerator:
    """流用元無しピッキングリスト自動生成クラス"""

    def __init__(self):
        # ===== フォルダパス設定 =====
        self.base_dir = Path(r"C:\temp\python_流用元無し_Pick")
        self.master_dir = Path(r"C:\temp\Newピッキング_対象照会\参照先")

        # ロガー
        self.logger = PickingLogger()

        # L/Rペアマップ
        self.lr_pairs = {
            "9R1": "9L1", "9L1": "9R1",
            "9R2A": "9L2A", "9L2A": "9R2A",
            "9R2B": "9L2B", "9L2B": "9R2B",
            "12R1": "12L1", "12L1": "12R1"
        }

        # 有効なレーン
        self.valid_lanes = {
            "7R1", "7R2", "7L1", "7L2",
            "9R1", "9R2A", "9R2B",
            "11R1", "11R2", "11R3", "11R4", "12R1",
            "9L1", "9L2A", "9L2B",
            "11L1", "11L2", "11L3", "12L1"
        }

    def clean_value(self, val):
        """値のクリーニング"""
        if pd.isna(val):
            return ""
        s = str(val).strip()
        if s.lower() == "nan" or s == "":
            return ""
        if s.isdigit() and len(s) <= 11:
            return int(s)
        return s

    # ============================================================================
    # ステップ1: ファイル検索
    # ============================================================================
    def find_target_file(self):
        """対象ファイルを検索"""
        print("=" * 80)
        print("【ステップ1】対象ファイル検索")
        print("=" * 80)

        print(f"📁 検索フォルダ: {self.base_dir}")

        all_files = list(self.base_dir.glob("*.xlsx"))
        all_files = [f for f in all_files if not f.name.startswith("~$")]

        if not all_files:
            raise FileNotFoundError(
                f"❌ 対象ファイルが見つかりません\n"
                f"フォルダ: {self.base_dir}\n"
                f"以下のフォルダに.xlsxファイルを配置してください"
            )

        valid_files = []
        for f in all_files:
            if "_Pick完了_" in f.stem:
                continue

            if len(f.stem) >= 10 and re.match(r'^[A-Za-z0-9]{10}', f.stem):
                valid_files.append(f)

        if len(valid_files) == 0:
            print(f"\n❌ 条件に合致するファイルがありません")
            print(f"\n見つかったファイル:")
            for i, f in enumerate(all_files, 1):
                print(f"  {i}. {f.name}")
            print(f"\n条件: 先頭10文字が英数字の.xlsxファイル（_Pick完了_を含まない）")
            raise FileNotFoundError("条件に合致するファイルが見つかりません")

        if len(valid_files) > 1:
            print("\n❌ エラー: 条件に合致するファイルが複数あります")
            print("\n見つかったファイル:")
            for i, f in enumerate(valid_files, 1):
                print(f"  {i}. {f.name}")
            print("\n1つのファイルのみにしてから再実行してください")
            raise ValueError("対象ファイルが複数存在します")

        target_file = valid_files[0]
        print(f"✅ 対象ファイル: {target_file.name}")

        self.logger.add_step("ファイル検索", f"対象: {target_file.name}")

        return target_file

    # ============================================================================
    # ステップ2: ファイル構造チェック
    # ============================================================================
    def validate_file_structure(self, file_path):
        """ファイル構造の妥当性チェック"""
        print("\n" + "=" * 80)
        print("【ステップ2】ファイル構造チェック")
        print("=" * 80)

        wb = load_workbook(file_path, read_only=True, data_only=True)

        if len(wb.sheetnames) != 1:
            raise ValueError(
                f"❌ シートが1枚ではありません\n"
                f"実際のシート数: {len(wb.sheetnames)}枚\n"
                f"シート名: {', '.join(wb.sheetnames)}"
            )

        print(f"✅ シート数: 1枚")

        ws = wb.active

        expected_headers = {
            'B1': '起点部品番号',
            'G1': 'レベル',
            'M1': '内外',
            'S1': 'FNC'
        }

        errors = []
        for cell_addr, expected_value in expected_headers.items():
            actual_value = str(ws[cell_addr].value or '').strip()

            if actual_value != expected_value:
                errors.append(f"  {cell_addr}: 期待='{expected_value}', 実際='{actual_value}'")
            else:
                print(f"✅ {cell_addr}: {expected_value}")

        wb.close()

        if errors:
            raise ValueError(
                f"❌ ヘッダーが正しくありません\n" + "\n".join(errors)
            )

        print(f"✅ ファイル構造OK")
        self.logger.add_step("ファイル構造チェック", "OK")

    # ============================================================================
    # ステップ3: 構成表マトリックス読込
    # ============================================================================
    def load_matrix(self, file_path):
        """構成表マトリックス読込"""
        print("\n" + "=" * 80)
        print("【ステップ3】構成表マトリックス読込")
        print("=" * 80)

        df = pd.read_excel(file_path, dtype=str).fillna("")
        df = df.map(self.clean_value)

        # ★★★ E1セル（Pythonのインデックスは0始まりなので、行=0, 列=4）★★★
        frame_number = str(df.iloc[0, 4]).strip()

        print(f"✅ フレーム品番（E1セル）: {frame_number}")
        print(f"✅ データ: {len(df)}行 × {len(df.columns)}列")

        # # デバッグ用に複数行表示
        # print(f"    E1セル: {df.iloc[0, 4]}")
        # print(f"    E2セル: {df.iloc[1, 4]}")
        # print(f"    E3セル: {df.iloc[2, 4]}")

        self.logger.add_step(
            "構成表マトリックス読込",
            f"フレーム品番: {frame_number}, 行数: {len(df)}"
        )

        return df, frame_number, file_path

    # ============================================================================
    # ステップ4: CMピッキング参照DB作成
    # ============================================================================
    def load_cm_master(self):
        """CMピッキングマスタ読込"""
        print("\n" + "=" * 80)
        print("【ステップ4】CMピッキング参照DB作成")
        print("=" * 80)

        cm_file = self.master_dir / "T_VJ_ＣＭピッキングマスタ.xlsx"
        if not cm_file.exists():
            raise FileNotFoundError(f"❌ CMマスタが見つかりません: {cm_file}")

        df_cm = pd.read_excel(cm_file, dtype=str).fillna("")

        data = []
        for _, row in df_cm.iterrows():
            for i in range(1, 21):
                lane = str(row.iloc[i]).strip()
                pattern = str(row.iloc[i + 20]).strip()

                if lane and pattern and pattern.lower() != "nan":
                    if pattern.startswith(('*', '5JX')) or not pattern[0].isdigit():
                        continue

                    data.append({
                        'レーン番号': i,
                        '出庫先レーン': lane,
                        '出庫先パターン': self.clean_value(pattern)
                    })

        df_ref = pd.DataFrame(data)

        counts = df_ref.groupby(['レーン番号', '出庫先レーン', '出庫先パターン']).size()
        valid_keys = counts[counts > 3].index

        df_ref = df_ref.set_index(['レーン番号', '出庫先レーン', '出庫先パターン'])
        df_ref = df_ref.loc[valid_keys].reset_index()

        print(f"✅ CMマスタ: {len(df_ref)}レコード")

        # ===== 【追加】除外表による行削除 ===== ★ここから追加
        df_ref = self._delete_rows_by_exclusion_list(df_ref)
        # ===== 【追加終了】 ===== ★ここまで追加

        self.logger.add_step("CMピッキング参照DB作成", f"{len(df_ref)}レコード")

        return df_ref

    # ============================================================================
    # ステップ5: A部品ピッキング参照DB作成
    # ============================================================================
    def load_parts_master(self):
        """部品出庫先テーブル読込"""
        print("\n" + "=" * 80)
        print("【ステップ5】A部品ピッキング参照DB作成")
        print("=" * 80)

        parts_file = self.master_dir / "T_VJ_部品出庫先テーブル.xlsx"
        if not parts_file.exists():
            raise FileNotFoundError(f"❌ 部品マスタが見つかりません: {parts_file}")

        df_parts = pd.read_excel(parts_file, dtype=str).fillna("")

        print(f"  元データの列数: {len(df_parts.columns)}列")

        # 最初の5列のみ抽出
        df_parts = df_parts.iloc[:, :5].copy()

        df_parts.columns = ['出庫先', 'パターン', '部品番号', '部品名称', 'pc']
        df_parts['出庫先'] = df_parts['出庫先'].str.strip().str.upper()

        df_ref = df_parts[
            ~df_parts['出庫先'].str.startswith(('CL', 'CR')) &
            df_parts['出庫先'].isin(self.valid_lanes)
            ].copy()

        for col in df_ref.columns:
            df_ref[col] = df_ref[col].apply(self.clean_value)

        # ===== 部品番号を文字列型に変換（構成表マトリックスとの一致のため） =====
        df_ref['部品番号'] = df_ref['部品番号'].astype(str).str.strip()


        # ソート
        df_ref = df_ref.sort_values(['出庫先', 'パターン']).reset_index(drop=True)

        # NO列と連番を設定（削除前）
        df_ref.insert(0, 'NO', range(1, len(df_ref) + 1))
        df_ref['連番'] = df_ref.groupby(['出庫先', 'パターン']).cumcount() + 1

        # ★★★ アンダーバー含む行を削除（連番設定後） ★★★
        before_count = len(df_ref)
        df_ref = df_ref[~df_ref['部品番号'].str.contains('_', na=False)].copy()
        after_count = len(df_ref)
        deleted_count = before_count - after_count

        if deleted_count > 0:
            print(f"      ✅ アンダーバー含む行を{deleted_count}件削除しました")
            self.logger.add_step("A部品参照DB_アンダーバー削除", f"{deleted_count}件削除")

        # NO列を振り直し（連番は振り直さない！）
        df_ref['NO'] = range(1, len(df_ref) + 1)

        df_ref['出庫先+パターン'] = df_ref['出庫先'] + df_ref['パターン'].astype(str)
        df_ref['チェック'] = '○'

        print(f"✅ 部品マスタ: {len(df_ref)}レコード")
        self.logger.add_step("A部品ピッキング参照DB作成", f"{len(df_ref)}レコード")

        return df_ref

    # ============================================================================
    # ステップ6: CMピッキング作成
    # ============================================================================
    def _insert_5th_cm_rows(self, df_cm, df_cm_ref):
        """4TH C/Mでレーン番号=2の行の下に5TH C/M行を挿入（VBA版準拠）"""
        print("\n  🔍 4TH C/M（レーン番号=2）の検索...")

        # ===== 行を複製して5TH C/M行を挿入 =====
        new_rows = []
        split_count = 0

        for idx, row in df_cm.iterrows():
            # 元の行を追加
            new_rows.append(row.to_dict())

            # 出庫先レーンが"4TH C/M"でレーン番号が"2"の場合
            lane_value = str(row.get('出庫先レーン', '')).strip()
            lane_number = str(row.get('レーン番号', '')).strip()

            if lane_value == "4TH C/M" and lane_number == "2":
                # 5TH C/M用の行を作成
                new_row = row.to_dict().copy()
                new_row['レーン番号'] = "5"
                new_row['出庫先レーン'] = "5TH C/M"
                new_rows.append(new_row)

                split_count += 1
                part_number = str(row.get('部品番号', '')).strip()
                print(f"      ✅ {part_number}: 4TH C/M → 5TH C/M行を挿入")

        # ===== DataFrameを作成 =====
        if len(new_rows) > 0:
            df_result = pd.DataFrame(new_rows)
            df_result = df_result.reset_index(drop=True)
        else:
            print(f"      ⚠️ 処理結果が空です")
            return df_cm

        # ===== 結果サマリー =====
        if split_count > 0:
            print(f"\n✅ 5TH C/M行を{split_count}件挿入しました")
            self.logger.add_step("5TH C/M行挿入", f"{split_count}件挿入")
        else:
            print(f"\n  ℹ️ 挿入対象なし（4TH C/Mでレーン番号2の行なし）")

        return df_result

    # ============================================================================
    # ステップ8: Excel保存
    # ============================================================================
    def save_to_excel(self, df_matrix, df_cm_ref, df_parts_ref, df_cm, df_a_picking,
                      frame_number, original_file, is_a_line):
        """Excel保存"""
        print("\n" + "=" * 80)
        print("【ステップ8】Excel保存")
        print("=" * 80)

        base_name = original_file.stem
        timestamp = datetime.now().strftime('%y%m%d_%H%M%S')  # ★★★ 秒まで追加 ★★★
        output_file = self.base_dir / f"{base_name}_Pick完了_{timestamp}.xlsx"

        print(f"📁 保存先: {self.base_dir}")

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_matrix.to_excel(writer, sheet_name='構成表マトリックス', index=False)
            df_cm_ref.to_excel(writer, sheet_name='CMピッキング参照DB', index=False)

            if is_a_line:
                df_parts_ref.to_excel(writer, sheet_name='A部品ピッキング参照DB', index=False)

            df_cm.to_excel(writer, sheet_name='CMピッキング', index=False)

            # 720システム入力シート作成
            self._create_720_input_sheet(writer, df_cm, df_a_picking, is_a_line)

            # CMピッキングデータを720システム入力に転記
            ws720 = writer.sheets['720システム入力']
            self._transfer_cm_picking_to_720(ws720, df_cm)

            # A部品ピッキングデータを720システム入力に転記
            if is_a_line:
                self._transfer_a_parts_picking_to_720(ws720, df_a_picking)

            # フレーム品番を720システム入力に転記
            self._transfer_frame_number_to_720(ws720, frame_number)

            # D/P情報を720システム入力に転記
            self._transfer_dp_info_to_720(ws720, df_matrix)

            # A部品ピッキングシートを作成
            if is_a_line:
                self._write_a_picking_sheet(writer, df_a_picking, df_matrix, df_parts_ref)

            # 720システム入力シートの最終書式設定
            self._format_720_system_sheet(ws720, is_a_line)

            # ★★★ 720システム入力シートを末尾に移動 ★★★
            wb = writer.book
            if '720システム入力' in wb.sheetnames:
                ws_720 = wb['720システム入力']
                wb.move_sheet(ws_720, offset=len(wb.sheetnames) - wb.index(ws_720) - 1)
                print("✅ 720システム入力シートを末尾に移動しました")

            # 列幅自動調整
            self._autofit_columns(writer.sheets['構成表マトリックス'])
            self._autofit_columns(writer.sheets['CMピッキング参照DB'])
            if is_a_line:
                self._autofit_columns(writer.sheets['A部品ピッキング参照DB'])

            # 枠線追加
            self._add_borders_to_sheet(writer.sheets['CMピッキング参照DB'], len(df_cm_ref))
            if is_a_line:
                self._add_borders_to_sheet(writer.sheets['A部品ピッキング参照DB'], len(df_parts_ref))

            # ★★★ CMピッキングシートの体裁整形 ★★★
            self._format_cm_sheet(writer.sheets['CMピッキング'], df_cm)

        print(f"✅ 保存完了: {output_file.name}")

        self.logger.add_step("ファイル保存", f"{output_file.name}")

        return output_file

    def _add_borders_to_sheet(self, ws, data_rows):
        """シート全体に枠線を追加"""
        thin = Side(border_style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # ヘッダー行とデータ行に枠線を追加
        max_col = ws.max_column
        for row in range(1, data_rows + 2):  # ヘッダー(1行目) + データ行
            for col in range(1, max_col + 1):
                ws.cell(row, col).border = border

    def _create_720_input_sheet(self, writer, df_cm, df_a_picking, is_a_line):
        """720システム入力シート作成"""
        ws = writer.book.create_sheet("720システム入力")

        thin = Side(border_style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # ========== CMピッキング ==========
        ws['A1'] = "CMピッキング"
        ws['A1'].font = Font(bold=True, size=11)

        # 新規:
        ws['B2'] = "新規："
        ws['B2'].font = Font(bold=True)

        cm_headers = [
            (3, "B", "(1)"), (3, "D", "(2)"), (3, "F", "(3)"),
            (3, "H", "(4)"), (3, "J", "(5)"),
            (7, "B", "(6)"), (7, "D", "(7)"), (7, "F", "(8)"),
            (7, "H", "(9)"), (7, "J", "(10)"),
            (11, "B", "(11)"), (11, "D", "(12)"), (11, "F", "(13)"),
            (11, "H", "(14)"), (11, "J", "(15)")
        ]

        for row, col, text in cm_headers:
            cell = ws[f"{col}{row}"]
            cell.value = text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 出庫先レーン・出庫先パターン
        ws['A4'] = "出庫先レーン"
        ws['A5'] = "出庫先パターン"
        ws['A8'] = "出庫先レーン"
        ws['A9'] = "出庫先パターン"
        ws['A12'] = "出庫先レーン"
        ws['A13'] = "出庫先パターン"

        # セル結合（B～K列の2列ずつ、行3,4,5,7,8,9,11,12,13）
        for row in [3, 4, 5, 7, 8, 9, 11, 12, 13]:
            for col in range(2, 12, 2):  # B,D,F,H,J列（2列ずつ）
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)

            # 出庫先レーン固定値（AラインとCラインで分岐）
            if is_a_line:
                # 出庫先レーン固定値_1_117_Aライン
                cm_lanes = [
                    (4, "B", "1TH C/M"), (4, "D", "2ND C/M"), (4, "F", "3RD C/M"), (4, "H", "4TH C/M"),
                    (4, "J", "5TH C/M"),
                    (8, "B", "6TH C/M"), (8, "D", "END C/M"), (8, "F", "RH  EXT"), (8, "H", "LH  EXT"),
                    (8, "J", "7TH C/M"),
                    (12, "B", "8TH C/M"), (12, "D", "ENG C/M"), (12, "H", "D/  PLT"), (12, "H", "RH C/BOX"),
                    (12, "J", "LH C/BOX")
                ]
            else:
                # 出庫先レーン固定値_2_113_Cライン
                cm_lanes = [
                    (4, "B", "1TH C/M"), (4, "D", "2ND C/M"), (4, "F", "3RD C/M"), (4, "H", "4TH C/M"),
                    (4, "J", "5TH C/M"),
                    (8, "B", "6TH C/M"), (8, "D", "END C/M"), (8, "F", "GUS 6C/M"), (8, "H", "T/M C/M"),
                    (8, "J", "RH EXTEN"),
                    (12, "B", "LH EXTEN"), (12, "D", "ENG C/M"), (12, "H", ""), (12, "H", ""), (12, "J", "")
                ]

        for row, col, text in cm_lanes:
            cell = ws[f"{col}{row}"]
            cell.value = text
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 出庫先パターンの入力欄（枠線のみ）
        cm_pattern_cells = [
            "B4", "C4", "D4", "E4", "F4", "G4", "H4", "I4", "J4", "K4",
            "B5", "C5", "D5", "E5", "F5", "G5", "H5", "I5", "J5", "K5",
            "B8", "C8", "D8", "E8", "F8", "G8", "H8", "I8", "J8", "K8",
            "B9", "C9", "D9", "E9", "F9", "G9", "H9", "I9", "J9", "K9",
            "B12", "C12", "D12", "E12", "F12", "G12", "H12", "I12", "J12", "K12",
            "B13", "C13", "D13", "E13", "F13", "G13", "H13", "I13", "J13", "K13",
        ]

        for cell_addr in cm_pattern_cells:
            ws[cell_addr].border = border

        for cell_addr in cm_pattern_cells:
            ws[cell_addr].border = border

        # ========== A部品ピッキング ==========
        ws['A18'] = "A部品ピッキング"
        ws['A18'].font = Font(bold=True, size=11)

        # 新規:
        ws['B19'] = "新規："
        ws['B19'].font = Font(bold=True)

        # レーン番号ヘッダー（1～20）
        lane_numbers = list(range(1, 21))
        for i, num in enumerate(lane_numbers[:10], start=2):  # B20～K20
            cell = ws.cell(20, i, num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # レーン名（7R1, 7R2, ...）
        lane_names_1 = ["7R1", "7R2", "7L1", "7L2", "9R1", "9R2A", "9R2B", "11R1", "11R2", "11R3"]
        for i, name in enumerate(lane_names_1, start=2):  # B21～K21
            cell = ws.cell(21, i, name)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 入力欄（B22～K22）
        for col in range(2, 12):
            ws.cell(22, col).border = border

        # 11～20
        for i, num in enumerate(lane_numbers[10:], start=2):  # B24～K24
            cell = ws.cell(24, i, num)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        lane_names_2 = ["11R4", "12R1", "9L1", "9L2A", "9L2B", "11L1", "11L2", "11L3", "12L1", ""]
        for i, name in enumerate(lane_names_2, start=2):  # B25～K25
            cell = ws.cell(25, i, name)
            if name:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 入力欄（B26～K26）
        for col in range(2, 12):
            ws.cell(26, col).border = border

        # 列幅設定
        ws.column_dimensions['A'].width = 14
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 12

    def _transfer_cm_picking_to_720(self, ws720, df_cm):
        """CMピッキング→720システム入力への転記"""
        print("\n  📋 CMピッキングデータを720システム入力に転記中...")

        transferred_count = 0

        # CMピッキングの各行を処理
        for idx, row in df_cm.iterrows():
            lane_name = str(row['出庫先レーン']).strip()

            # 出庫先レーンが空白でない場合のみ処理
            if lane_name and lane_name != "":
                lane_number = str(row['レーン番号']).strip()
                pattern = str(row['部品番号']).strip()

                # レーン番号を検索
                target_cell = None
                for row in [3, 7, 11]:
                    for col in range(2, 12):
                        cell_value = str(ws720.cell(row, col).value or "").strip()
                        if cell_value.replace("(", "").replace(")", "") == lane_number:
                            target_cell = (row, col)
                            break
                    if target_cell:
                        break

                # 見つかった場合は転記
                if target_cell:
                    base_row, base_col = target_cell
                    ws720.cell(base_row + 1, base_col, lane_name)  # 出庫先レーン
                    ws720.cell(base_row + 2, base_col, pattern)  # パターン
                    transferred_count += 1

        print(f"      ✅ {transferred_count}件転記完了")
        self.logger.add_step("CMピッキング→720転記", f"{transferred_count}件")

    def _transfer_a_parts_picking_to_720(self, ws720, df_a_picking):
        """A部品ピッキング→720システム入力への転記"""
        print("\n  📋 A部品ピッキングデータを720システム入力に転記中...")

        transferred_count = 0

        # A部品ピッキングの各行を処理
        for idx, row in df_a_picking.iterrows():
            lane_name = str(row['出庫先']).strip()
            pattern = str(row['パターン']).strip()

            # 出庫先とパターンが空白でない場合のみ処理
            if lane_name and pattern and lane_name != "" and pattern != "":
                # 720シートから該当レーン名を検索（B21:K21, B25:K25の範囲）
                target_cell = None

                # 1～10: B21:K21
                for col in range(2, 12):  # B～K列
                    cell_value = str(ws720.cell(21, col).value or "").strip()
                    if cell_value == lane_name:
                        target_cell = (21, col)
                        break

                # 11～20: B25:K25
                if not target_cell:
                    for col in range(2, 12):
                        cell_value = str(ws720.cell(25, col).value or "").strip()
                        if cell_value == lane_name:
                            target_cell = (25, col)
                            break

                # 見つかった場合は転記
                if target_cell:
                    base_row, base_col = target_cell

                    # パターンを2桁表示に変換
                    if pattern.isdigit():
                        pattern = pattern.zfill(2)  # 例: "5" → "05"

                    ws720.cell(base_row + 1, base_col, pattern)  # パターン
                    ws720.cell(base_row + 1, base_col).number_format = '@'  # 文字列形式
                    transferred_count += 1

        print(f"      ✅ {transferred_count}件転記完了")
        self.logger.add_step("A部品ピッキング→720転記", f"{transferred_count}件")

    def _transfer_frame_number_to_720(self, ws720, frame_number):
        """フレーム品番を720システム入力に転記"""
        print("\n  📋 フレーム品番を720システム入力に転記中...")

        ws720['C2'] = frame_number
        ws720['C19'] = frame_number

        print(f"      ✅ フレーム品番: {frame_number}")
        self.logger.add_step("フレーム品番→720転記", f"{frame_number}")

    def _format_720_system_sheet(self, ws720, is_a_line):
        """720システム入力シートの最終書式設定"""
        print("\n  🎨 720システム入力シートの書式設定中...")

        # C1セルにライン表示
        if is_a_line:
            ws720['C1'] = "Aライン_117"
        else:
            ws720['C1'] = "Cライン_113"

        ws720['C1'].font = Font(name="Meiryo UI", bold=True, size=11)
        ws720['C1'].alignment = Alignment(horizontal='left', vertical='center')

        # Cラインの場合、A部品ピッキング部分（18～26行）を非表示
        if not is_a_line:
            for row_num in range(18, 27):  # 18～26行
                ws720.row_dimensions[row_num].hidden = True

        # フォント変更
        for row in ws720.iter_rows():
            for cell in row:
                cell.font = Font(name="Meiryo UI", size=cell.font.size, bold=cell.font.bold)

        # 「新規:」と品番の配置調整
        ws720['B2'].alignment = Alignment(horizontal='right', vertical='center')
        ws720['C2'].alignment = Alignment(horizontal='left', vertical='center')
        ws720['B19'].alignment = Alignment(horizontal='right', vertical='center')
        ws720['C19'].alignment = Alignment(horizontal='left', vertical='center')

        # レーン番号をカッコ付き表示に変換
        lane_number_ranges = [
            "B3", "D3", "F3", "H3", "J3",  # (1)～(5)
            "B7", "D7", "F7", "H7", "J7",  # (6)～(10)
            "B11", "D11", "F11", "H11", "J11"  # (11)～(15)
        ]

        thin = Side(border_style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for cell_addr in lane_number_ranges:
            cell = ws720[cell_addr]
            cell_value = str(cell.value or "").strip()

            if cell_value and cell_value.isdigit():
                cell.value = f"({cell_value})"
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 罫線の再設定（全体枠）
        for row_start in [3, 7, 11]:
            for row in range(row_start, row_start + 3):
                for col in range(2, 12):  # B～K列
                    ws720.cell(row, col).border = border

        # A部品ピッキングの罫線
        for row_start in [20, 24]:
            for row in range(row_start, row_start + 3):
                for col in range(2, 12):
                    ws720.cell(row, col).border = border

        # データ入力行の中央寄せ（5, 9, 13, 22, 26行目のB～K列）
        data_rows = [5, 9, 13, 22, 26]
        for row in data_rows:
            for col in range(2, 12):  # B～K列
                cell = ws720.cell(row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # フレーム品番を太字に設定
        ws720['C2'].font = Font(name="Meiryo UI", bold=True)
        ws720['C19'].font = Font(name="Meiryo UI", bold=True)

        print(f"      ✅ 書式設定完了")
        self.logger.add_step("720システム書式設定", "完了")

    def _transfer_dp_info_to_720(self, ws720, df_matrix):
        """D/P情報を720システム入力に転記"""
        print("\n  📋 D/P情報を検索中...")

        # 外部ファイルパス
        source_file = self.master_dir / "New_部品ピッキング諸元.xlsx"

        if not source_file.exists():
            print(f"      ⚠️ D/Pマスタファイルが見つかりません: {source_file}")
            return

        try:
            # 外部ファイル読込
            df_source = pd.read_excel(source_file, dtype=str).fillna("")

            # D列（インデックス3）が"#D/P"の行を抽出
            if len(df_source.columns) < 4:
                print(f"      ⚠️ D/Pマスタの列数が不足しています")
                return

            df_dp = df_source[df_source.iloc[:, 3].astype(str).str.strip() == "#D/P"].copy()

            if len(df_dp) == 0:
                print(f"      ℹ️ D/P情報が見つかりません")
                return

            # 構成表マトリックスE列から品番を取得
            if len(df_matrix.columns) < 5:
                print(f"      ⚠️ 構成表マトリックスの列数が不足しています")
                return

            e_col = df_matrix.columns[4]

            found = False
            for idx in df_matrix.index:
                part_number = str(df_matrix.loc[idx, e_col]).strip()

                if len(part_number) >= 9:
                    prefix = part_number[:9]

                    # D/PマスタのA列と部分一致を検索
                    for dp_idx in df_dp.index:
                        if len(df_source.columns) > 0:
                            dp_part = str(df_dp.loc[dp_idx, df_dp.columns[0]]).strip()

                            if prefix in dp_part:
                                dp_value = str(df_dp.loc[dp_idx, df_dp.columns[3]]).strip()
                                ws720['F12'] = dp_value
                                ws720['F13'] = "*"
                                ws720['F13'].alignment = Alignment(horizontal='center', vertical='center')

                                print(f"      ✅ D/P情報を転記: {dp_value}")
                                self.logger.add_step("D/P情報→720転記", f"{dp_value}")
                                found = True
                                break

                if found:
                    break

            if not found:
                print(f"      ℹ️ D/P情報は該当なし")

        except Exception as e:
            print(f"      ⚠️ D/P情報転記エラー: {e}")
            import traceback
            traceback.print_exc()

    # ============================================================================
    # ステップ6: CMピッキング作成
    # ============================================================================
    def create_cm_picking(self, df_matrix, df_cm_ref):
        """CMピッキング作成"""
        print("\n" + "=" * 80)
        print("【ステップ6】CMピッキング作成")
        print("=" * 80)

        g_col = df_matrix.columns[6]
        df_cm = df_matrix[df_matrix[g_col] == '_1'].copy()

        df_cm = df_cm.iloc[:, [4, 5, 9]].copy()
        df_cm.columns = ['部品番号', '部品名称', '数量']

        df_cm['部品番号'] = df_cm['部品番号'].astype(str).str.replace('*', '', regex=False).str.strip()

        df_cm['レーン番号'] = ""
        df_cm['出庫先レーン'] = ""

        # CMピッキング参照DBから部品番号ごとにレーン情報を取得
        ref_dict = {}
        for _, row in df_cm_ref.iterrows():
            pattern = str(row['出庫先パターン']).strip()
            if pattern not in ref_dict:
                ref_dict[pattern] = {
                    'lane_num': str(row['レーン番号']),
                    'lane': row['出庫先レーン']
                }

        # CMピッキングの各部品に対してレーン情報を割り当て
        yellow_count = 0
        for idx in df_cm.index:
            part_num = df_cm.loc[idx, '部品番号']

            # CMピッキング参照DBのC列（出庫先パターン）から部品番号を検索
            if part_num in ref_dict:
                df_cm.loc[idx, 'レーン番号'] = ref_dict[part_num]['lane_num']
                df_cm.loc[idx, '出庫先レーン'] = ref_dict[part_num]['lane']
                yellow_count += 1

        print(f"✅ CMピッキング: {len(df_cm)}部品")
        print(f"✅ 黄色ハイライト対象: {yellow_count}部品")

        self.logger.add_step(
            "CMピッキング作成",
            f"総数: {len(df_cm)}部品, 黄色ハイライト: {yellow_count}部品"
        )

        # 5TH C/M行の挿入（4TH C/Mでレーン番号が2の行を複製）
        df_cm = self._insert_5th_cm_rows(df_cm, df_cm_ref)

        return df_cm

    # ============================================================================
    # ステップ7: A部品ピッキング作成（VBA版完全準拠）
    # ============================================================================

    def _assign_patterns_simple(self, df_result, df_parts_ref):
        """パターン割り当て（数量無視・高速版）（VBA: AssignPatternToE_Method3_Simple）"""

        # A~E列を出庫先順にソート
        df_result.sort_values('出庫先', inplace=True)
        df_result.reset_index(drop=True, inplace=True)

        # L/Rペアマップ
        lr_pair_map = self.lr_pairs

        # 参照DBを辞書化（部品番号のみ）
        pattern_dict = {}
        for _, row in df_parts_ref.iterrows():
            lane = str(row['出庫先']).strip()
            pattern = str(row['パターン']).strip()
            part = str(row['部品番号']).strip()

            if lane and pattern and part:
                key = f"{lane}|{pattern}"
                if key not in pattern_dict:
                    pattern_dict[key] = set()
                pattern_dict[key].add(part)

        # 出庫先ごとにパターンを割り当て
        current_lane = ""
        required_parts = set()
        required_quantities = {}
        start_idx = 0

        for idx in range(len(df_result) + 1):
            if idx < len(df_result):
                lane = str(df_result.loc[idx, '出庫先']).strip()
            else:
                lane = ""

            if lane != current_lane:
                if current_lane and required_parts:
                    # 最適パターンを検索
                    best_pattern = self._find_best_pattern_simple(
                        current_lane, required_parts, required_quantities,
                        pattern_dict, lr_pair_map
                    )

                    # パターンを設定
                    for i in range(start_idx, idx):
                        if str(df_result.loc[i, '出庫先']).strip() == current_lane:
                            if best_pattern:
                                df_result.loc[i, 'パターン'] = best_pattern
                            else:
                                df_result.loc[i, 'パターン'] = "合致パターン無し"

                if idx < len(df_result):
                    current_lane = lane
                    part = df_result.loc[idx, '部品番号']
                    qty = df_result.loc[idx, 'PC']

                    required_parts = {part}
                    required_quantities = {part: qty}
                    start_idx = idx
            else:
                if idx < len(df_result):
                    part = df_result.loc[idx, '部品番号']
                    qty = df_result.loc[idx, 'PC']

                    required_parts.add(part)
                    required_quantities[part] = required_quantities.get(part, 0) + qty

        failed_count = (df_result['パターン'] == "合致パターン無し").sum()
        print(f"      結果: 失敗 {failed_count}行")

    def _autofit_columns(self, ws):
        """列幅を自動調整"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # 最大50文字まで
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_cm_sheet(self, ws, df_cm):
        """CMピッキングシート書式設定（シンプル版）"""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        thin = Side(border_style="thin")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 列幅設定
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

        # ヘッダー行を太字・中央寄せ
        for col in range(1, 6):
            cell = ws.cell(1, col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # データ行の書式設定
        for row in range(2, len(df_cm) + 2):
            for col in range(1, 6):
                cell = ws.cell(row, col)
                cell.border = border

                # D列を中央寄せ
                if col == 4:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # D列とE列に値がある行を黄色に
            d_value = ws.cell(row, 4).value
            e_value = ws.cell(row, 5).value
            if d_value and e_value and str(d_value).strip() != "" and str(e_value).strip() != "":
                for col in range(1, 6):
                    ws.cell(row, col).fill = yellow_fill

    def create_a_picking_list(self, df_matrix, df_parts_ref):
        """A部品ピッキングリスト作成（VBA完全移植版）"""
        print("\n" + "=" * 80)
        print("【ステップ7】A部品ピッキングリスト作成（VBA完全移植版）")
        print("=" * 80)

        # ===== ステップ1: 構成表マトリックスからレベル="_1"の行を抽出 =====
        print("\n  7-1. 構成表マトリックスからレベル_1を抽出")
        g_col = df_matrix.columns[6]
        df_target = df_matrix[df_matrix[g_col] == '_1'].copy()

        # E列(部品番号)、F列(部品名称)、J列(数量)を抽出
        df_target = df_target.iloc[:, [4, 5, 9]].copy()
        df_target.columns = ['部品番号', '部品名称', 'PC']

        # 部品番号のクリーニング（*削除）
        df_target['部品番号'] = df_target['部品番号'].astype(str).str.replace('*', '', regex=False).str.strip()

        # 数値変換
        df_target['PC'] = pd.to_numeric(df_target['PC'], errors='coerce').fillna(1).astype(int)

        print(f"      必要部品: {len(df_target)}行")

        # ===== ステップ2: 参照DBに存在する部品のみ残す =====
        print("\n  7-2. 参照DBに存在する部品のみ抽出")

        # 参照DBの部品番号リスト
        ref_parts = set(df_parts_ref['部品番号'].unique())

        # 該当部品のみ残す
        df_target = df_target[df_target['部品番号'].isin(ref_parts)].copy()

        print(f"      抽出後: {len(df_target)}行")

        # ===== ステップ3: D列・E列を追加（空白） =====
        df_target['出庫先'] = ""
        df_target['パターン'] = ""

        # 列順を調整: A=部品番号, B=部品名称, C=PC, D=出庫先, E=パターン
        df_target = df_target[['部品番号', '部品名称', 'PC', '出庫先', 'パターン']].copy()

        # ===== ステップ4: G~M列に参照DB情報を転記 =====
        print("\n  7-3. 参照DBからG~M列に全データを転記")

        # 部品番号ごとに参照DBの行を抽出
        result_rows = []

        for idx, row in df_target.iterrows():
            part_num = row['部品番号']

            # 該当部品の全行を参照DBから取得
            matching_rows = df_parts_ref[df_parts_ref['部品番号'] == part_num].copy()

            if len(matching_rows) > 0:
                # 各マッチング行を結果に追加
                for _, ref_row in matching_rows.iterrows():
                    result_rows.append({
                        '部品番号': row['部品番号'],
                        '部品名称': row['部品名称'],
                        'PC': row['PC'],
                        '出庫先': "",
                        'パターン': "",
                        'G_出庫先': ref_row['出庫先'],
                        'H_パターン': ref_row['パターン'],
                        'I_連番': ref_row.get('連番', ""),
                        'J_出庫先+パターン': ref_row.get('出庫先+パターン', ""),
                        'K_部品番号': ref_row['部品番号'],
                        'L_部品名称': ref_row['部品名称'],
                        'M_pc': ref_row['pc']
                    })

        df_result = pd.DataFrame(result_rows)

        print(f"      転記後: {len(df_result)}行")

        # ===== ステップ5: G~M列をソート =====
        print("\n  7-4. G~M列をソート（出庫先→パターン→連番）")

        if len(df_result) > 0:
            # パターンを数値化してソート
            df_result['H_パターン_数値'] = pd.to_numeric(df_result['H_パターン'], errors='coerce')
            df_result['I_連番_数値'] = pd.to_numeric(df_result['I_連番'], errors='coerce')

            df_result = df_result.sort_values(
                ['G_出庫先', 'H_パターン_数値', 'I_連番_数値']
            ).reset_index(drop=True)

            df_result = df_result.drop(columns=['H_パターン_数値', 'I_連番_数値'])

        # ===== ステップ6: 候補リスト作成 =====
        print("\n  7-5. 候補リスト作成")
        self._create_candidate_list(df_result)

        # ===== ステップ7: 出庫先割り当て =====
        print("\n  7-6. 出庫先割り当て")
        self._assign_lanes_from_candidates(df_result)

        # ===== ステップ8: パターン割り当て =====
        print("\n  7-7. パターン割り当て（高速版）")
        self._assign_patterns_simple(df_result, df_parts_ref)

        # ===== ステップ9: 失敗したケースのみ再試行 =====
        print("\n  7-8. 失敗ケースの再試行（精密版）")
        self._retry_failed_patterns(df_result, df_parts_ref)

        # ===== A~E列のみ残す =====
        df_final = df_result[['部品番号', '部品名称', 'PC', '出庫先', 'パターン']].copy()

        # ===== 結果サマリー =====
        total = len(df_final)
        unique_parts = df_final['部品番号'].nunique()
        unique_lanes = df_final['出庫先'].nunique()

        print(f"\n  ✅ 処理完了:")
        print(f"      総行数: {total}行")
        print(f"      ユニーク部品数: {unique_parts}種類")
        print(f"      出庫先数: {unique_lanes}箇所")

        self.logger.add_step(
            "A部品ピッキング作成",
            f"総数: {total}行, 部品種類: {unique_parts}, 出庫先: {unique_lanes}"
        )

        return df_final

    def _create_candidate_list(self, df_result):
        """候補リストをO~R列に作成（VBA: CreateLanePatternCandidateList）"""
        # G列・K列からユニークな組み合わせを抽出
        candidates_data = []
        seen = set()

        for idx, row in df_result.iterrows():
            part = str(row['K_部品番号']).strip()
            lane = str(row['G_出庫先']).strip()

            if part and lane:
                key = f"{part}|{lane}"
                if key not in seen:
                    candidates_data.append({
                        'O_部品番号': part,
                        'P_出庫先': lane,
                        'Q_パターン': str(row['H_パターン']).strip(),
                        'R_連番': str(row['I_連番']).strip()
                    })
                    seen.add(key)

        # ソート
        candidates_df = pd.DataFrame(candidates_data)
        if len(candidates_df) > 0:
            candidates_df = candidates_df.sort_values(['O_部品番号', 'P_出庫先']).reset_index(drop=True)

            # 結果に列を追加
            for col in ['O_部品番号', 'P_出庫先', 'Q_パターン', 'R_連番']:
                df_result[col] = ""

            # データを転記
            for i, row in candidates_df.iterrows():
                if i < len(df_result):
                    for col, val in row.items():
                        df_result.loc[i, col] = val

        print(f"      候補リスト: {len(candidates_data)}件")

    def _assign_lanes_from_candidates(self, df_result):
        """出庫先を候補リストから割り当て（VBA: AssignLaneFromCandidateList）"""

        # 数量2以上の行を分割
        print("      数量2以上の行を分割中...")
        self._split_quantity_rows(df_result)

        # L/Rペアマップ
        lr_pair_map = self.lr_pairs

        # 候補情報を辞書化
        candidates_dict = {}

        # O~R列から候補情報を読み取る
        for idx, row in df_result.iterrows():
            part = str(row.get('O_部品番号', '')).strip()
            lane = str(row.get('P_出庫先', '')).strip()
            pattern = str(row.get('Q_パターン', '')).strip()

            if part and lane:
                if part not in candidates_dict:
                    candidates_dict[part] = []

                # 重複チェック
                if not any(c['lane'] == lane for c in candidates_dict[part]):
                    candidates_dict[part].append({
                        'lane': lane,
                        'pattern': pattern
                    })

        # 使用済み出庫先を記録
        used_lanes = {}

        # 各行に出庫先を割り当て
        assigned_count = 0
        for idx in df_result.index:
            if df_result.loc[idx, '出庫先'] == "":
                part = df_result.loc[idx, '部品番号']

                if part not in used_lanes:
                    used_lanes[part] = []

                # 候補を取得
                candidates = candidates_dict.get(part, [])

                if len(candidates) == 1:
                    # 候補が1つの場合は確定
                    selected = candidates[0]
                    df_result.loc[idx, '出庫先'] = selected['lane']
                    used_lanes[part].append(selected['lane'])
                    assigned_count += 1

                elif len(candidates) >= 2:
                    # 未使用の候補を抽出
                    available = [
                        c for c in candidates
                        if c['lane'] not in used_lanes[part]
                    ]

                    if not available:
                        available = candidates

                    if available:
                        # 最初の候補を使用
                        selected = available[0]
                        df_result.loc[idx, '出庫先'] = selected['lane']
                        used_lanes[part].append(selected['lane'])
                        assigned_count += 1

        print(f"      出庫先割り当て: {assigned_count}行")

    def _split_quantity_rows(self, df_result):
        """数量2以上の行を分割（VBA: SplitQuantity2Rows）"""
        new_rows = []

        for idx, row in df_result.iterrows():
            qty = row['PC']

            if qty >= 2:
                # 数量1の行を複数作成
                for _ in range(qty):
                    new_row = row.copy()
                    new_row['PC'] = 1
                    new_row['出庫先'] = ""
                    new_rows.append(new_row)
            else:
                new_rows.append(row)

        # データフレームを置き換え
        df_result.drop(df_result.index, inplace=True)
        for i, row_data in enumerate(new_rows):
            for col in df_result.columns:
                df_result.loc[i, col] = row_data[col]

    def _assign_patterns_simple(self, df_result, df_parts_ref):
        """パターン割り当て（数量無視・高速版）（VBA: AssignPatternToE_Method3_Simple）"""

        # A~E列を出庫先順にソート
        df_result.sort_values('出庫先', inplace=True)
        df_result.reset_index(drop=True, inplace=True)

        # L/Rペアマップ
        lr_pair_map = self.lr_pairs

        # 参照DBを辞書化（部品番号のみ）
        pattern_dict = {}
        for _, row in df_parts_ref.iterrows():
            lane = str(row['出庫先']).strip()
            pattern = str(row['パターン']).strip()
            part = str(row['部品番号']).strip()

            if lane and pattern and part:
                key = f"{lane}|{pattern}"
                if key not in pattern_dict:
                    pattern_dict[key] = set()
                pattern_dict[key].add(part)

        # 出庫先ごとにパターンを割り当て
        current_lane = ""
        required_parts = set()
        start_idx = 0

        for idx in range(len(df_result) + 1):
            if idx < len(df_result):
                lane = str(df_result.loc[idx, '出庫先']).strip()
            else:
                lane = ""

            if lane != current_lane:
                if current_lane and required_parts:
                    # 最適パターンを検索
                    best_pattern = self._find_best_pattern_simple(
                        current_lane, required_parts, pattern_dict, lr_pair_map
                    )

                    # パターンを設定
                    for i in range(start_idx, idx):
                        if str(df_result.loc[i, '出庫先']).strip() == current_lane:
                            if best_pattern:
                                df_result.loc[i, 'パターン'] = best_pattern
                            else:
                                df_result.loc[i, 'パターン'] = "合致パターン無し"

                if idx < len(df_result):
                    current_lane = lane
                    part = df_result.loc[idx, '部品番号']
                    required_parts = {part}
                    start_idx = idx
            else:
                if idx < len(df_result):
                    part = df_result.loc[idx, '部品番号']
                    required_parts.add(part)

        failed_count = (df_result['パターン'] == "合致パターン無し").sum()
        print(f"      結果: 失敗 {failed_count}行")

    def _find_best_pattern_simple(self, lane, required_parts, pattern_dict, lr_pair_map):
        """最適パターンを検索（数量無視版）"""

        # L/R専用部品を検出
        lr_specific_parts = {}
        lr_prefix_groups = {}
        common_parts = set()

        for part1 in required_parts:
            if len(part1) >= 10:
                prefix1 = part1[:8]

                for part2 in required_parts:
                    if part1 != part2 and len(part2) >= 10:
                        if part2[:8] == prefix1 and part1[-2:] != part2[-2:]:
                            if prefix1 not in lr_prefix_groups:
                                lr_prefix_groups[prefix1] = []

                            if part1 not in lr_prefix_groups[prefix1]:
                                lr_prefix_groups[prefix1].append(part1)
                                lr_specific_parts[part1] = prefix1

        # ペアの出庫先を取得
        pair_lane = lr_pair_map.get(lane, "")

        # 最適パターンを検索
        best_pattern = ""
        best_score = (-1, 999999, 999999)

        for key, pattern_parts in pattern_dict.items():
            if not key.startswith(f"{lane}|"):
                continue

            pattern = key.split("|")[1]

            match_count = 0
            missing_count = 0

            for part in required_parts:
                matched = False

                # 共通部品の場合
                if part in common_parts:
                    if part in pattern_parts:
                        match_count += 1
                        matched = True

                # L/R専用部品の場合
                elif part in lr_specific_parts:
                    if part in pattern_parts:
                        match_count += 1
                        matched = True
                    elif pair_lane:
                        # ペアの出庫先に該当部品があるかチェック
                        pair_key = f"{pair_lane}|{pattern}"
                        if pair_key in pattern_dict:
                            prefix = lr_specific_parts[part]
                            for pair_part in lr_prefix_groups.get(prefix, []):
                                if pair_part != part and pair_part in pattern_dict[pair_key]:
                                    match_count += 1
                                    matched = True
                                    break

                # その他の部品
                else:
                    if part in pattern_parts:
                        match_count += 1
                        matched = True

                if not matched:
                    missing_count += 1

            excess_count = len(pattern_parts) - match_count
            score = (match_count, -missing_count, -excess_count)

            if score > best_score:
                best_score = score
                best_pattern = pattern

        return best_pattern

    def _retry_failed_patterns(self, df_result, df_parts_ref):
        """失敗したパターンを再試行（数量考慮版）（VBA: RetryFailedLanesWithQuantityMatching）"""

        # 失敗した出庫先を抽出
        failed_lanes = df_result[df_result['パターン'] == "合致パターン無し"]['出庫先'].unique()

        if len(failed_lanes) == 0:
            print(f"      再試行不要: 全て成功")
            return

        print(f"      再試行対象: {len(failed_lanes)}出庫先")

        # L/Rペアマップ
        lr_pair_map = self.lr_pairs

        # 参照DBを辞書化（数量込み）
        pattern_dict_qty = {}
        for _, row in df_parts_ref.iterrows():
            lane = str(row['出庫先']).strip()
            pattern = str(row['パターン']).strip()
            part = str(row['部品番号']).strip()
            pc = str(row['pc']).strip()

            if lane and pattern and part and pc:
                key = f"{lane}|{pattern}"
                if key not in pattern_dict_qty:
                    pattern_dict_qty[key] = set()
                pattern_dict_qty[key].add(f"{part}|{pc}")

        # 失敗した出庫先ごとに再試行
        for lane in failed_lanes:
            group = df_result[df_result['出庫先'] == lane]

            # 部品と数量を集計
            part_qty = {}
            for _, row in group.iterrows():
                part = row['部品番号']
                qty = row['PC']
                part_qty[part] = part_qty.get(part, 0) + qty

            # 最適パターンを検索（数量考慮版）
            best_pattern = self._find_best_pattern_with_qty(
                lane, part_qty, pattern_dict_qty, lr_pair_map
            )

            if best_pattern:
                df_result.loc[group.index, 'パターン'] = best_pattern
                print(f"        ✅ {lane}: パターン{best_pattern}を割り当て")
            else:
                print(f"        ❌ {lane}: パターンが見つかりません")

    def _find_best_pattern_with_qty(self, lane, part_qty, pattern_dict_qty, lr_pair_map):
        """最適パターン検索（数量考慮版）（VBA: FindBestPattern）"""

        # L/R専用部品を検出
        lr_specific_parts = {}
        lr_prefix_groups = {}

        for part1 in part_qty.keys():
            if len(part1) >= 10:
                prefix1 = part1[:8]

                for part2 in part_qty.keys():
                    if part1 != part2 and len(part2) >= 10:
                        if part2[:8] == prefix1 and part1[-2:] != part2[-2:]:
                            if prefix1 not in lr_prefix_groups:
                                lr_prefix_groups[prefix1] = []

                            if part1 not in lr_prefix_groups[prefix1]:
                                lr_prefix_groups[prefix1].append(part1)
                                lr_specific_parts[part1] = prefix1

        # ペアの出庫先を取得
        pair_lane = lr_pair_map.get(lane, "")

        # 最適パターンを検索
        best_pattern = ""
        best_match_count = -1
        best_missing_count = 999999
        best_excess_count = 999999

        for key, pattern_parts in pattern_dict_qty.items():
            key_parts = key.split("|")
            cand_lane = key_parts[0]
            cand_pattern = key_parts[1]

            # 同じ出庫先のみ評価
            if cand_lane != lane:
                continue

            match_count = 0
            missing_count = 0

            # 必要な部品ごとに一致判定（数量込み）
            for part, qty in part_qty.items():
                part_matched = False
                check_key = f"{part}|{qty}"

                # L/R専用部品の場合
                if part in lr_specific_parts:
                    if check_key in pattern_parts:
                        match_count += 1
                        part_matched = True
                    elif pair_lane:
                        # ペアの出庫先に該当部品があるかチェック
                        pair_key = f"{pair_lane}|{cand_pattern}"

                        if pair_key in pattern_dict_qty:
                            prefix = lr_specific_parts[part]

                            for pair_part in lr_prefix_groups.get(prefix, []):
                                pair_check_key = f"{pair_part}|{qty}"

                                if pair_part != part and pair_check_key in pattern_dict_qty[pair_key]:
                                    match_count += 1
                                    part_matched = True
                                    break

                # その他の部品
                else:
                    if check_key in pattern_parts:
                        match_count += 1
                        part_matched = True

                if not part_matched:
                    missing_count += 1

            # 余剰数
            excess_count = len(pattern_parts) - match_count

            # 最適パターン判定
            is_better = False

            if match_count > best_match_count:
                is_better = True
            elif match_count == best_match_count:
                if missing_count < best_missing_count:
                    is_better = True
                elif missing_count == best_missing_count and excess_count < best_excess_count:
                    is_better = True

            if is_better:
                best_pattern = cand_pattern
                best_match_count = match_count
                best_missing_count = missing_count
                best_excess_count = excess_count

        # 不足がある場合は採用しない（VBA準拠）
        if best_missing_count > 0:
            return ""

        return best_pattern

    def _is_lr_pair(self, lane1, lane2):
        """2つのレーンがL/Rペアか判定（VBA: IsLRPair）"""
        # 長さチェック
        if len(lane1) != len(lane2):
            return False

        if len(lane1) < 2:
            return False

        # 2文字目がLとRの組み合わせかチェック
        char1 = lane1[1]
        char2 = lane2[1]

        is_lr = (char1 == 'L' and char2 == 'R') or (char1 == 'R' and char2 == 'L')

        if not is_lr:
            return False

        # 2文字目以外が一致するかチェック
        rest1 = lane1[0] + lane1[2:]
        rest2 = lane2[0] + lane2[2:]

        return rest1 == rest2

    def _write_a_picking_sheet(self, writer, df_a_picking, df_matrix, df_parts_ref):
        """A部品ピッキングシート書き込み（VBA完全準拠版）"""
        ws = writer.book.create_sheet("A部品ピッキング")
        ws720 = writer.sheets['720システム入力']

        # ★★★ 色・線の定義（メソッドの最初で定義） ★★★
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

        data_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="medium", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        thin = Side(border_style="thin")
        thick = Side(border_style="medium")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        # ========================================================================
        # ステップ1: df_a_pickingから出庫先+パターンの組み合わせを取得
        # ========================================================================
        print("\n  📋 df_a_pickingから出庫先+パターンを抽出中...")

        # df_a_pickingの出庫先+パターンのユニークな組み合わせ
        unique_combinations = df_a_picking[['出庫先', 'パターン']].drop_duplicates()

        print(f"      出庫先+パターンの組み合わせ: {len(unique_combinations)}件")

        # ========================================================================
        # ステップ2: 各出庫先・パターンごとに参照DBから全部品を抽出
        # ========================================================================
        picking_list_rows = []

        for _, comb in unique_combinations.iterrows():
            lane = comb['出庫先']
            pattern = str(comb['パターン']).strip()

            if not lane or not pattern:
                continue

            # 参照DBから該当する出庫先・パターンの全部品を抽出
            matched_parts = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ].copy()

            for _, row in matched_parts.iterrows():
                picking_list_rows.append({
                    '部品番号': row['部品番号'],
                    '部品名称': row['部品名称'],
                    'PC': row['pc'],
                    '出庫先': lane,
                    'パターン': pattern
                })

        df_picking_list = pd.DataFrame(picking_list_rows)

        # アンダーバー（_）を含む部品番号の行を削除
        before_count = len(df_picking_list)
        df_picking_list = df_picking_list[~df_picking_list['部品番号'].astype(str).str.contains('_', na=False)].copy()
        after_count = len(df_picking_list)
        deleted_count = before_count - after_count

        if deleted_count > 0:
            print(f"      ✅ アンダーバー含む行を{deleted_count}件削除しました")

        print(f"      ピッキングリスト作成: {len(df_picking_list)}行")

        # ========================================================================
        # 左上: [ピッキングリスト確定版] A1～
        # ========================================================================
        ws['A1'] = "[ピッキングリスト確定版]"
        ws['A1'].font = Font(bold=True, size=12)

        headers_picking = ['部品番号', '部品名称', 'PC', '出庫先', 'パターン']
        for col_idx, header in enumerate(headers_picking, start=1):
            cell = ws.cell(2, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # データ書き込み（ソート→書き込み）
        df_picking_list = df_picking_list.sort_values(['出庫先', 'パターン']).reset_index(drop=True)

        row_idx = 3
        for idx, row_data in df_picking_list.iterrows():
            for col_idx, value in enumerate([
                row_data['部品番号'],
                row_data['部品名称'],
                row_data['PC'],
                row_data['出庫先'],
                row_data['パターン']
            ], start=1):
                ws.cell(row_idx, col_idx, value)
            row_idx += 1

        picking_end_row = row_idx - 1



        picking_end_row = row_idx - 1

        # ========================================================================
        # 右側: [出庫先+パターン 出庫先特定用_ALL] G1～
        # ========================================================================
        ws['G1'] = "[出庫先+パターン 出庫先特定用_ALL]"
        ws['G1'].font = Font(bold=True, size=12)

        headers_ref = ['出庫先', 'パターン', '連番', '出庫先+パターン', '部品番号', '部品名称', 'pc']
        for col_idx, header in enumerate(headers_ref, start=7):
            cell = ws.cell(2, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 候補データ作成
        print("\n  📋 候補データ作成中...")

        try:
            # 構成表マトリックスのE列から必要部品番号を抽出
            g_col = df_matrix.columns[6]
            df_required = df_matrix[df_matrix[g_col] == '_1'].copy()
            required_part_numbers = set(
                df_required.iloc[:, 4].astype(str).str.replace('*', '', regex=False).str.strip()
            )

            # アンダーバー含む部品は除外
            required_part_numbers = {p for p in required_part_numbers if '_' not in p and p != ''}

            print(f"      必要部品数: {len(required_part_numbers)}件")

            # 候補データ作成
            df_candidates = df_parts_ref[
                df_parts_ref['部品番号'].isin(required_part_numbers)
            ].copy()

            if '連番' not in df_candidates.columns:
                df_candidates['連番'] = ''

            # ソート
            df_candidates['パターン_数値'] = pd.to_numeric(df_candidates['パターン'], errors='coerce')
            df_candidates = df_candidates.sort_values(
                ['出庫先', 'パターン_数値', '部品番号']
            ).reset_index(drop=True)

            print(f"      候補データ: {len(df_candidates)}件抽出")

        except Exception as e:
            print(f"      ⚠️ 候補データ作成エラー: {e}")
            import traceback
            traceback.print_exc()
            df_candidates = pd.DataFrame(columns=['出庫先', 'パターン', '連番', '部品番号', '部品名称', 'pc'])

        # 候補データをG～M列に表示
        row_idx = 3
        try:
            for _, row_data in df_candidates.iterrows():
                ws.cell(row_idx, 7, str(row_data.get('出庫先', ''))).border = border
                ws.cell(row_idx, 8, str(row_data.get('パターン', ''))).border = border
                ws.cell(row_idx, 9, str(row_data.get('連番', ''))).border = border
                ws.cell(row_idx, 10, f"{row_data.get('出庫先', '')}{row_data.get('パターン', '')}").border = border
                ws.cell(row_idx, 11, str(row_data.get('部品番号', ''))).border = border
                ws.cell(row_idx, 12, str(row_data.get('部品名称', ''))).border = border
                ws.cell(row_idx, 13, str(row_data.get('pc', ''))).border = border
                row_idx += 1

            print(f"      ✅ 候補マスタ（G～M列）: {len(df_candidates)}件表示")

        except Exception as e:
            print(f"      ⚠️ 候補マスタ書き込みエラー: {e}")
            import traceback
            traceback.print_exc()

        # ========================================================================
        # 左下: [ピッキングリスト検証用] A(N+3)～
        # ========================================================================
        verification_start_row = picking_end_row + 3
        ws.cell(verification_start_row, 1, "[ピッキングリスト検証用_A部品ピッキング一覧から抽出]")
        ws.cell(verification_start_row, 1).font = Font(bold=True, size=12)

        headers_verification = ['部品番号', '部品名称', 'PC', '一致']
        for col_idx, header in enumerate(headers_verification, start=1):
            cell = ws.cell(verification_start_row + 1, col_idx, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 構成表マトリックスから必要部品を抽出
        g_col = df_matrix.columns[6]
        df_target = df_matrix[df_matrix[g_col] == '_1'].copy()
        df_target = df_target.iloc[:, [4, 5, 9]].copy()
        df_target.columns = ['部品番号', '部品名称', 'PC']
        df_target['部品番号'] = df_target['部品番号'].astype(str).str.replace('*', '', regex=False).str.strip()
        df_target['PC'] = pd.to_numeric(df_target['PC'], errors='coerce').fillna(1).astype(int)

        # アンダーバー含む部品を除外
        df_target = df_target[~df_target['部品番号'].astype(str).str.contains('_', na=False)].copy()

        # 左上の表からA列（部品番号）ごとにC列（PC）を集計
        picking_summary = df_picking_list.groupby('部品番号')['PC'].sum().to_dict()

        # A部品ピッキング対象のみ抽出
        df_target = df_target[df_target['部品番号'].isin(picking_summary.keys())].copy()

        # 部品番号・部品名称ごとにPCを集計
        df_verification = df_target.groupby(['部品番号', '部品名称'], as_index=False)['PC'].sum()

        df_verification = df_verification.sort_values('部品番号').reset_index(drop=True)

        # 判定列を追加
        df_verification['一致'] = df_verification.apply(
            lambda row: '○' if picking_summary.get(row['部品番号'], 0) == row['PC'] else '×',
            axis=1
        )

        # ×が1つでもあれば再試行
        mismatch_count = (df_verification['一致'] == '×').sum()
        if mismatch_count > 0:
            mismatch_parts = df_verification[df_verification['一致'] == '×']['部品番号'].tolist()
            print(f"\n⚠️ A部品ピッキング数量不一致: {mismatch_count}件")
            print(f"   不一致部品: {', '.join(mismatch_parts[:10])}")
            if len(mismatch_parts) > 10:
                print(f"   ...他{len(mismatch_parts) - 10}件")

            # # ===== 改善版再試行処理 =====
            # print(f"\n  🔄 改善版再試行を開始...")
            # retry_success = self._retry_with_smart_pattern_search(
            #     df_picking_list, df_verification, df_candidates, df_parts_ref, ws720, df_matrix
            # )

            # ===== 大改修版：最適パターン組み合わせ探索 =====
            print(f"\n  🔄 最適パターン組み合わせ探索を開始...")

            # optimal_combination = self._find_optimal_pattern_combination(
            #     mismatch_parts, df_verification, df_parts_ref, df_picking_list, df_matrix
            # )
            #
            # retry_success = optimal_combination is not None

            print(f"      🔍 デバッグ: mismatch_parts = {mismatch_parts}")
            print(f"      🔍 デバッグ: df_picking_list の行数 = {len(df_picking_list)}")

            optimal_combination = self._find_optimal_pattern_combination(
                mismatch_parts, df_verification, df_parts_ref, df_picking_list, df_matrix
            )

            print(f"      🔍 デバッグ: optimal_combination = {optimal_combination}")

            retry_success = optimal_combination is not None
            print(f"      🔍 デバッグ: retry_success = {retry_success}")


            if retry_success:
                # print(f"      ✅ 再試行により不一致を解消しました")

                print(f"      ✅ 最適パターン組み合わせを発見しました")

                # ★★★ 既存の不一致部品の行を全て削除 ★★★
                print(f"      🗑️ 既存の不一致部品行を削除中...")
                df_picking_list = df_picking_list[~df_picking_list['部品番号'].isin(mismatch_parts)].copy()
                df_picking_list = df_picking_list.reset_index(drop=True)

                # # ★★★ 最適パターンを追加 ★★★
                # print(f"      ➕ 最適パターンを追加中...")
                # for part, patterns in optimal_combination.items():
                #     for pattern_info in patterns:
                #         # このパターンの全部品を参照DBから取得
                #         pattern_parts = df_parts_ref[
                #             (df_parts_ref['出庫先'] == pattern_info['lane']) &
                #             (df_parts_ref['パターン'].astype(str) == pattern_info['pattern'])
                #             ]
                #
                #         for _, part_row in pattern_parts.iterrows():
                #             new_row = {
                #                 '部品番号': part_row['部品番号'],
                #                 '部品名称': part_row['部品名称'],
                #                 'PC': part_row['pc'],
                #                 '出庫先': pattern_info['lane'],
                #                 'パターン': pattern_info['pattern']
                #             }
                #             df_picking_list.loc[len(df_picking_list)] = new_row

                # ★★★ 最適パターンを追加（必要な部品のみ） ★★★
                print(f"      ➕ 最適パターンを追加中...")
                for part, patterns in optimal_combination.items():
                    for pattern_info in patterns:
                        # 必要な部品の行のみを参照DBから取得
                        part_row = df_parts_ref[
                            (df_parts_ref['出庫先'] == pattern_info['lane']) &
                            (df_parts_ref['パターン'].astype(str) == pattern_info['pattern']) &
                            (df_parts_ref['部品番号'] == part)
                            ]

                        if len(part_row) > 0:
                            part_row = part_row.iloc[0]
                            new_row = {
                                '部品番号': part_row['部品番号'],
                                '部品名称': part_row['部品名称'],
                                'PC': part_row['pc'],
                                '出庫先': pattern_info['lane'],
                                'パターン': pattern_info['pattern']
                            }
                            df_picking_list.loc[len(df_picking_list)] = new_row
                            print(f"         追加: {part} → {pattern_info['lane']}:{pattern_info['pattern']}")



                print(f"      ✅ 最適パターン反映完了")




                # ★★★ 左上のピッキングリストを再書き込み ★★★
                print(f"      📝 ピッキングリストを再書き込み中...")

                # 既存の行をクリア（3行目から50行程度）
                for clear_row in range(3, 3 + 100):
                    for clear_col in range(1, 6):
                        ws.cell(clear_row, clear_col).value = None

                # 更新されたdf_picking_listを再書き込み
                row_idx = 3
                for idx, row_data in df_picking_list.iterrows():
                    for col_idx, value in enumerate([
                        row_data['部品番号'],
                        row_data['部品名称'],
                        row_data['PC'],
                        row_data['出庫先'],
                        row_data['パターン']
                    ], start=1):
                        cell = ws.cell(row_idx, col_idx, value)
                        cell.border = border
                    row_idx += 1

                picking_end_row = row_idx - 1
                print(f"      ✅ ピッキングリスト再書き込み完了: {len(df_picking_list)}行")

                # ★★★ 再書き込み後のグレー背景処理 ★★★
                print(f"      🎨 グループごとのグレー背景を再設定中...")

                group_start = 3
                color_index = 0

                df_picking_list_sorted = df_picking_list.sort_values(['出庫先', 'パターン']).reset_index(drop=True)

                for idx in range(len(df_picking_list_sorted)):
                    is_last = (idx == len(df_picking_list_sorted) - 1)

                    if not is_last:
                        curr_lane = df_picking_list_sorted.iloc[idx]['出庫先']
                        curr_pat = df_picking_list_sorted.iloc[idx]['パターン']
                        next_lane = df_picking_list_sorted.iloc[idx + 1]['出庫先']
                        next_pat = df_picking_list_sorted.iloc[idx + 1]['パターン']
                        changed = (curr_lane != next_lane) or (curr_pat != next_pat)
                    else:
                        changed = True

                    if changed:
                        group_end = idx + 3

                        # グレー背景（1つおき）
                        if color_index == 1:
                            for row in range(group_start, group_end + 1):
                                for col in range(1, 6):
                                    ws.cell(row, col).fill = data_gray

                        # 太枠
                        for col in range(1, 6):
                            ws.cell(group_start, col).border = Border(
                                top=thick,
                                left=thick if col == 1 else ws.cell(group_start, col).border.left,
                                right=thick if col == 5 else ws.cell(group_start, col).border.right,
                                bottom=ws.cell(group_start, col).border.bottom
                            )
                            ws.cell(group_end, col).border = Border(
                                top=ws.cell(group_end, col).border.top,
                                left=thick if col == 1 else ws.cell(group_end, col).border.left,
                                right=thick if col == 5 else ws.cell(group_end, col).border.right,
                                bottom=thick
                            )

                        for row in range(group_start, group_end + 1):
                            ws.cell(row, 1).border = Border(
                                top=thick if row == group_start else ws.cell(row, 1).border.top,
                                left=thick,
                                right=ws.cell(row, 1).border.right,
                                bottom=thick if row == group_end else ws.cell(row, 1).border.bottom
                            )
                            ws.cell(row, 5).border = Border(
                                top=thick if row == group_start else ws.cell(row, 5).border.top,
                                left=ws.cell(row, 5).border.left,
                                right=thick,
                                bottom=thick if row == group_end else ws.cell(row, 5).border.bottom
                            )

                        group_start = group_end + 1
                        color_index = 1 - color_index

                print(f"      ✅ グレー背景設定完了")

                # 検証表の位置を更新
                verification_start_row = picking_end_row + 3

                # 検証表のタイトルとヘッダーを再配置
                ws.cell(verification_start_row, 1, "[ピッキングリスト検証用_A部品ピッキング一覧から抽出]")
                ws.cell(verification_start_row, 1).font = Font(bold=True, size=12)

                headers_verification = ['部品番号', '部品名称', 'PC', '一致']
                for col_idx, header in enumerate(headers_verification, start=1):
                    cell = ws.cell(verification_start_row + 1, col_idx, header)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

                # 検証表を再作成
                picking_summary = df_picking_list.groupby('部品番号')['PC'].sum().to_dict()
                df_verification['一致'] = df_verification.apply(
                    lambda row: '○' if picking_summary.get(row['部品番号'], 0) == row['PC'] else '×',
                    axis=1
                )

                mismatch_count = (df_verification['一致'] == '×').sum()
                if mismatch_count > 0:
                    print(f"      ⚠️ 再試行後も不一致: {mismatch_count}件")
                else:
                    print(f"      ✅ 全ての不一致を解消しました")
            else:
                print(f"      ⚠️ 再試行失敗: 代替パターンが見つかりませんでした")

            self.logger.add_warning("A部品ピッキング数量不一致", f"{mismatch_count}件")

        # データ書き込み
        row_idx = verification_start_row + 2
        for _, row_data in df_verification.iterrows():
            ws.cell(row_idx, 1, row_data['部品番号']).border = border
            ws.cell(row_idx, 2, row_data['部品名称']).border = border
            ws.cell(row_idx, 3, row_data['PC']).border = border

            # D列（一致）をセンタリング
            cell_d = ws.cell(row_idx, 4, row_data['一致'])
            cell_d.border = border
            cell_d.alignment = Alignment(horizontal='center', vertical='center')

            row_idx += 1

        verification_end_row = row_idx - 1

        # ========================================================================
        # 左上ピッキングリストの書式設定
        # ========================================================================
        print("\n  🎨 ピッキングリストの書式設定中...")

        # ★★★ グループごとのグレー背景処理 ★★★
        # 色・線の定義
        # data_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # thin = Side(border_style="thin", color="000000")
        # thick = Side(border_style="medium", color="000000")

        group_start = 3
        color_index = 0

        # df_picking_listをソート
        df_picking_list_sorted = df_picking_list.sort_values(['出庫先', 'パターン']).reset_index(drop=True)

        for idx in range(len(df_picking_list_sorted)):
            is_last = (idx == len(df_picking_list_sorted) - 1)

            if not is_last:
                curr_lane = df_picking_list_sorted.iloc[idx]['出庫先']
                curr_pat = str(df_picking_list_sorted.iloc[idx]['パターン'])
                next_lane = df_picking_list_sorted.iloc[idx + 1]['出庫先']
                next_pat = str(df_picking_list_sorted.iloc[idx + 1]['パターン'])
                changed = (curr_lane != next_lane) or (curr_pat != next_pat)
            else:
                changed = True

            if changed:
                group_end = idx + 3

                # グレー背景（1つおき）
                if color_index == 1:
                    for row in range(group_start, group_end + 1):
                        for col in range(1, 6):
                            ws.cell(row, col).fill = data_gray

                # 太枠
                for col in range(1, 6):
                    cell_top = ws.cell(group_start, col)
                    cell_top.border = Border(
                        top=thick,
                        left=thick if col == 1 else thin,
                        right=thick if col == 5 else thin,
                        bottom=cell_top.border.bottom
                    )

                    cell_bottom = ws.cell(group_end, col)
                    cell_bottom.border = Border(
                        top=cell_bottom.border.top,
                        left=thick if col == 1 else thin,
                        right=thick if col == 5 else thin,
                        bottom=thick
                    )

                for row in range(group_start, group_end + 1):
                    cell_left = ws.cell(row, 1)
                    cell_left.border = Border(
                        top=thick if row == group_start else thin,
                        left=thick,
                        right=thin,
                        bottom=thick if row == group_end else thin
                    )

                    cell_right = ws.cell(row, 5)
                    cell_right.border = Border(
                        top=thick if row == group_start else thin,
                        left=thin,
                        right=thick,
                        bottom=thick if row == group_end else thin
                    )

                group_start = group_end + 1
                color_index = 1 - color_index

        # 色・線の定義
        header_gray = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        # data_gray = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        # thin = Side(border_style="thin", color="000000")
        # thick = Side(border_style="medium", color="000000")

        # ヘッダー行（A2:E2）
        for col in range(1, 6):
            cell = ws.cell(2, col)
            cell.fill = header_gray
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                top=thick,
                left=thick if col == 1 else thin,
                right=thick if col == 5 else thin,
                bottom=thick
            )

        # データ行の格子線
        for row in range(3, picking_end_row + 1):
            for col in range(1, 6):
                cell = ws.cell(row, col)
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if col == 5:  # E列を中央揃え
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # 外枠（左・右・下）
        for row in range(3, picking_end_row + 1):
            ws.cell(row, 1).border = Border(top=thin, left=thick, right=thin, bottom=thin)
            ws.cell(row, 5).border = Border(top=thin, left=thin, right=thick, bottom=thin)

        for col in range(1, 6):
            cell = ws.cell(picking_end_row, col)
            cell.border = Border(
                top=thin,
                left=thick if col == 1 else thin,
                right=thick if col == 5 else thin,
                bottom=thick
            )

            # 出庫先+パターンの塊ごとに太枠とグレー背景
            group_start = 3
            color_index = 0

            # インデックスをリセット（0から連番に）
            df_picking_list = df_picking_list.reset_index(drop=True)

            for idx in range(len(df_picking_list)):
                is_last = (idx == len(df_picking_list) - 1)

                if not is_last:
                    curr_lane = df_picking_list.iloc[idx]['出庫先']
                    curr_pat = df_picking_list.iloc[idx]['パターン']
                    next_lane = df_picking_list.iloc[idx + 1]['出庫先']
                    next_pat = df_picking_list.iloc[idx + 1]['パターン']
                    changed = (curr_lane != next_lane) or (curr_pat != next_pat)
                else:
                    changed = True

            if changed:
                group_end = idx + 3

                # グレー背景（1つおき）
                if color_index == 1:
                    for row in range(group_start, group_end + 1):
                        for col in range(1, 6):
                            ws.cell(row, col).fill = data_gray

                # 太枠
                for col in range(1, 6):
                    ws.cell(group_start, col).border = Border(
                        top=thick,
                        left=thick if col == 1 else ws.cell(group_start, col).border.left,
                        right=thick if col == 5 else ws.cell(group_start, col).border.right,
                        bottom=ws.cell(group_start, col).border.bottom
                    )
                    ws.cell(group_end, col).border = Border(
                        top=ws.cell(group_end, col).border.top,
                        left=thick if col == 1 else ws.cell(group_end, col).border.left,
                        right=thick if col == 5 else ws.cell(group_end, col).border.right,
                        bottom=thick
                    )

                for row in range(group_start, group_end + 1):
                    ws.cell(row, 1).border = Border(
                        top=thick if row == group_start else ws.cell(row, 1).border.top,
                        left=thick,
                        right=ws.cell(row, 1).border.right,
                        bottom=thick if row == group_end else ws.cell(row, 1).border.bottom
                    )
                    ws.cell(row, 5).border = Border(
                        top=thick if row == group_start else ws.cell(row, 5).border.top,
                        left=ws.cell(row, 5).border.left,
                        right=thick,
                        bottom=thick if row == group_end else ws.cell(row, 5).border.bottom
                    )

                group_start = group_end + 1
                color_index = 1 - color_index

        print(f"      ✅ 書式設定完了")

        # 列幅設定
        ws.column_dimensions['A'].width = 15

        # 列幅設定
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 3
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 5
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 35
        ws.column_dimensions['M'].width = 5

        # ========================================================================
        # ★★★ 720システム入力シートを更新 ★★★
        # ========================================================================
        print("\n  📝 720システム入力シートを更新中...")

        # A部品ピッキングの最新データを720シートに転記
        # まず全てクリア
        for row in [22, 26]:
            for col in range(2, 12):
                ws720.cell(row, col).value = ""

        # df_picking_listから出庫先+パターンのユニークな組み合わせを取得
        unique_lanes = df_picking_list[['出庫先', 'パターン']].drop_duplicates()

        transferred_count = 0

        for _, row_data in unique_lanes.iterrows():
            lane = row_data['出庫先']
            pattern = str(row_data['パターン']).strip()

            if not lane or not pattern:
                continue

            # 720シートから該当レーン名を検索（B21:K21, B25:K25）
            target_cell = None

            for row in [21, 25]:
                for col in range(2, 12):
                    cell_value = str(ws720.cell(row, col).value or "").strip()
                    if cell_value == lane:
                        target_cell = (row, col)
                        break
                if target_cell:
                    break

            if target_cell:
                row_num, col_num = target_cell

                # パターンを2桁表示に変換
                if pattern.isdigit():
                    pattern = pattern.zfill(2)

                ws720.cell(row_num + 1, col_num).value = pattern
                ws720.cell(row_num + 1, col_num).number_format = '@'
                transferred_count += 1

        print(f"      ✅ 720シート更新完了: {transferred_count}件")

    def _retry_with_smart_pattern_search(self, df_picking_list, df_verification, df_candidates,
                                         df_parts_ref, ws720, df_matrix):
        """
        改善版再試行処理：孤立部品の代替パターンを全出庫先から探索

        【処理フロー】
        1. 不一致部品を検出
        2. 出庫先ごとに数量を集計
        3. L/Rペアで割り当て可能な場合は優先
        4. 余剰の出庫先から孤立部品を検出
        5. 全出庫先から孤立部品のみのパターンを探索
        6. 見つかったらdf_picking_listと720シートを更新
        """

        # L/Rペアマップ
        lr_pair_map = self.lr_pairs

        # 不一致部品を抽出
        failed_parts = df_verification[df_verification['一致'] == '×']['部品番号'].tolist()

        if not failed_parts:
            return False

        print(f"         不一致部品: {len(failed_parts)}件")
        for part in failed_parts:
            required_qty = df_verification[df_verification['部品番号'] == part]['PC'].iloc[0]
            print(f"           - {part}: {required_qty}個必要")

        # 参照DBをパターン辞書に変換
        pattern_dict = {}
        for _, row in df_parts_ref.iterrows():
            lane = str(row['出庫先']).strip()
            pattern = str(row['パターン']).strip()
            part = str(row['部品番号']).strip()
            pc = int(row['pc']) if pd.notna(row['pc']) else 1

            key = f"{lane}|{pattern}"
            if key not in pattern_dict:
                pattern_dict[key] = {}
            pattern_dict[key][part] = pc

        success_count = 0

        # 不一致部品ごとに処理
        for failed_part in failed_parts:
            required_qty = df_verification[df_verification['部品番号'] == failed_part]['PC'].iloc[0]

            # 現在の割り当て状況を取得
            current_assignments = df_picking_list[df_picking_list['部品番号'] == failed_part]

            print(f"\n         部品 {failed_part} ({required_qty}個) の処理:")
            print(f"            現在の割り当て: {len(current_assignments)}行")
            for _, row in current_assignments.iterrows():
                print(f"              - {row['出庫先']}:{row['パターン']} ({row['PC']}個)")

            # 出庫先ごとの数量集計
            lane_qty = current_assignments.groupby('出庫先')['PC'].sum().to_dict()

            # ===== ケース1: 数量2でL/Rペアが使える場合 =====
            if required_qty == 2 and len(lane_qty) >= 2:
                # L/Rペアを探す
                lanes = list(lane_qty.keys())
                lr_pair_found = None

                for i, lane1 in enumerate(lanes):
                    for lane2 in lanes[i + 1:]:
                        if self._is_lr_pair(lane1, lane2):
                            lr_pair_found = (lane1, lane2)
                            break
                    if lr_pair_found:
                        break

                if lr_pair_found:
                    lane1, lane2 = lr_pair_found
                    print(f"            ✅ L/Rペア発見: {lane1} + {lane2}")

                    # # 削除する出庫先を特定
                    # removed_lanes = [lane for lane in lane_qty.keys() if lane not in [lane1, lane2]]


                    # # 削除する出庫先を特定（既に削除済みの出庫先は除外）
                    # removed_lanes = []
                    # for lane in lane_qty.keys():
                    #     if lane not in [lane1, lane2]:
                    #         # この出庫先がまだdf_picking_listに存在するか確認
                    #         if len(df_picking_list[df_picking_list['出庫先'] == lane]) > 0:
                    #             removed_lanes.append(lane)
                    #         else:
                    #             print(f"            {lane} は既に削除済み（スキップ）")

                    # 削除する出庫先を特定（既に削除済みの出庫先は除外）
                    print(f"            🔍 デバッグ: lane_qty.keys() = {list(lane_qty.keys())}")
                    print(f"            🔍 デバッグ: L/Rペア = {lane1}, {lane2}")
                    removed_lanes = []
                    for lane in lane_qty.keys():
                        print(f"            🔍 デバッグ: {lane} をチェック中...")
                        if lane not in [lane1, lane2]:
                            # この出庫先がまだdf_picking_listに存在するか確認
                            lane_exists = len(df_picking_list[df_picking_list['出庫先'] == lane]) > 0
                            print(f"            🔍 デバッグ: {lane} は存在する? {lane_exists}")
                            if lane_exists:
                                removed_lanes.append(lane)
                            else:
                                print(f"            {lane} は既に削除済み（スキップ）")
                    print(f"            🔍 デバッグ: removed_lanes = {removed_lanes}")






                    # ★★★ 削除する出庫先の全部品を削除し、孤立部品を検出 ★★★
                    for removed_lane in removed_lanes:
                        # この出庫先のパターンを取得

                        removed_rows = df_picking_list[df_picking_list['出庫先'] == removed_lane]

                        if len(removed_rows) > 0:
                            removed_pattern = removed_rows['パターン'].iloc[0]


                            print(f"            削除対象: {removed_lane}:{removed_pattern}")

                            # このパターンの全部品を参照DBから取得
                            removed_parts = df_parts_ref[
                                (df_parts_ref['出庫先'] == removed_lane) &
                                (df_parts_ref['パターン'].astype(str) == str(removed_pattern))
                                ]['部品番号'].tolist()

                            print(f"              削除部品: {removed_parts}")

                            # # これらの部品を全て削除
                            # df_picking_list.drop(
                            #     df_picking_list[
                            #         (df_picking_list['出庫先'] == removed_lane) &
                            #         (df_picking_list['パターン'].astype(str) == str(removed_pattern))
                            #         ].index,
                            #     inplace=True
                            # )
                            #
                            # # 削除した部品の中から孤立部品を検出
                            # isolated_parts = set(removed_parts) - {failed_part}

                            # これらの部品を全て削除
                            df_picking_list.drop(
                                df_picking_list[
                                    (df_picking_list['出庫先'] == removed_lane) &
                                    (df_picking_list['パターン'].astype(str) == str(removed_pattern))
                                    ].index,
                                inplace=True
                            )

                            # ★★★ DataFrameのインデックスをリセット ★★★
                            df_picking_list = df_picking_list.reset_index(drop=True)

                            # 削除した部品の中から孤立部品を検出
                            isolated_parts = set(removed_parts) - {failed_part}







                            if len(isolated_parts) > 0:
                                print(f"            孤立部品検出: {isolated_parts}")

                                # ★★★ 孤立部品を数量別にグループ化 ★★★
                                parts_qty_2 = set()  # 数量2個の部品
                                parts_qty_3_4 = set()  # 数量3-4個の部品
                                parts_qty_1 = set()  # 数量1個の部品

                                for isolated_part in isolated_parts:
                                    # 必要数量を取得（構成表マトリックスから）
                                    part_qty_rows = df_matrix[
                                        (df_matrix.iloc[:, 6] == '_1') &
                                        (df_matrix.iloc[:, 4].astype(str).str.replace('*', '',
                                                                                      regex=False).str.strip() == isolated_part)
                                        ]

                                    if len(part_qty_rows) > 0:
                                        required_qty = pd.to_numeric(part_qty_rows.iloc[0, 9], errors='coerce')
                                        if pd.notna(required_qty):
                                            required_qty = int(required_qty)

                                            if required_qty == 2:
                                                parts_qty_2.add(isolated_part)
                                            elif required_qty in [3, 4]:
                                                parts_qty_3_4.add(isolated_part)
                                            else:
                                                parts_qty_1.add(isolated_part)
                                        else:
                                            parts_qty_1.add(isolated_part)
                                    else:
                                        parts_qty_1.add(isolated_part)

                                print(f"              数量2個: {len(parts_qty_2)}部品")
                                print(f"              数量3-4個: {len(parts_qty_3_4)}部品")
                                print(f"              数量1個: {len(parts_qty_1)}部品")

                                alternative_found = None

                                # ★★★ 優先度1: 数量2個 → 9L2A + 9R2A ペア ★★★
                                if len(parts_qty_2) > 0:
                                    print(f"            優先探索: 数量2個部品を 9L2A + 9R2A ペアで探索...")

                                    # 9L2Aのパターンを探す
                                    pattern_9l2a = self._find_pattern_containing_parts_with_qty(
                                        '9L2A', parts_qty_2, 1, df_parts_ref, df_picking_list
                                    )

                                    # 9R2Aのパターンを探す
                                    pattern_9r2a = self._find_pattern_containing_parts_with_qty(
                                        '9R2A', parts_qty_2, 1, df_parts_ref, df_picking_list
                                    )

                                    if pattern_9l2a and pattern_9r2a:
                                        alternative_found = ('qty2_pair', '9L2A', pattern_9l2a, '9R2A', pattern_9r2a,
                                                             parts_qty_2)
                                        print(
                                            f"            ✅ 数量2個ペア発見: 9L2A:{pattern_9l2a} + 9R2A:{pattern_9r2a}")

                                # ★★★ 優先度2: 数量3-4個 → 7L2 + 7R2 ペア ★★★
                                if not alternative_found and len(parts_qty_3_4) > 0:
                                    print(f"            次善策: 数量3-4個部品を 7L2 + 7R2 ペアで探索...")

                                    # 7L2のパターンを探す
                                    pattern_7l2 = self._find_pattern_containing_parts_any_qty(
                                        '7L2', parts_qty_3_4, df_parts_ref, df_picking_list
                                    )

                                    # 7R2のパターンを探す
                                    pattern_7r2 = self._find_pattern_containing_parts_any_qty(
                                        '7R2', parts_qty_3_4, df_parts_ref, df_picking_list
                                    )

                                    if pattern_7l2 and pattern_7r2:
                                        alternative_found = ('qty3_4_pair', '7L2', pattern_7l2, '7R2', pattern_7r2,
                                                             parts_qty_3_4)
                                        print(f"            ✅ 数量3-4個ペア発見: 7L2:{pattern_7l2} + 7R2:{pattern_7r2}")

                                # ★★★ 優先度3: 数量1個 → 単独パターン ★★★
                                if not alternative_found and len(parts_qty_1) > 0:
                                    print(f"            次善策: 数量1個部品を単独パターンで探索...")

                                    # 全出庫先から探索
                                    for _, cand_row in df_parts_ref.iterrows():
                                        alt_lane = cand_row['出庫先']
                                        alt_pattern = str(cand_row['パターン']).strip()

                                        pattern_rows = df_parts_ref[
                                            (df_parts_ref['出庫先'] == alt_lane) &
                                            (df_parts_ref['パターン'].astype(str) == alt_pattern)
                                            ]

                                        pattern_parts = set(pattern_rows['部品番号'].tolist())

                                        # 数量1個の部品全てが含まれているかチェック
                                        if parts_qty_1.issubset(pattern_parts):
                                            # 未使用かチェック
                                            already_used = df_picking_list[
                                                (df_picking_list['出庫先'] == alt_lane) &
                                                (df_picking_list['パターン'].astype(str) == alt_pattern)
                                                ]

                                            if len(already_used) == 0:
                                                alternative_found = ('qty1_single', alt_lane, alt_pattern, parts_qty_1)
                                                print(f"            ✅ 数量1個パターン発見: {alt_lane}:{alt_pattern}")
                                                break

                                # ★★★ 混合ケース: 数量2個 + 数量1個 ★★★
                                if not alternative_found and len(parts_qty_2) > 0 and len(parts_qty_1) > 0:
                                    print(f"            次善策: 数量2個+1個を混合パターンで探索...")

                                    # 9L2A側: 数量2個(各1個) + 数量1個の一部
                                    # 9R2A側: 数量2個(各1個) + 数量1個の残り

                                    # すべての数量1個部品の分割パターンを試す
                                    from itertools import combinations
                                    parts_1_list = list(parts_qty_1)

                                    for split_count in range(len(parts_1_list) + 1):
                                        for group1_parts in combinations(parts_1_list, split_count):
                                            group1 = parts_qty_2 | set(group1_parts)
                                            group2 = parts_qty_2 | (parts_qty_1 - set(group1_parts))

                                            # 9L2Aで group1 を探す（数量2個は1個ずつ）
                                            pattern_9l2a = self._find_mixed_pattern(
                                                '9L2A', parts_qty_2, 1, set(group1_parts), 1, df_parts_ref,
                                                df_picking_list
                                            )

                                            # 9R2Aで group2 を探す（数量2個は1個ずつ）
                                            pattern_9r2a = self._find_mixed_pattern(
                                                '9R2A', parts_qty_2, 1, parts_qty_1 - set(group1_parts), 1,
                                                df_parts_ref, df_picking_list
                                            )

                                            if pattern_9l2a and pattern_9r2a:
                                                alternative_found = ('mixed', '9L2A', pattern_9l2a, '9R2A',
                                                                     pattern_9r2a, parts_qty_2 | parts_qty_1)
                                                print(
                                                    f"            ✅ 混合パターン発見: 9L2A:{pattern_9l2a} + 9R2A:{pattern_9r2a}")
                                                break

                                        if alternative_found:
                                            break

                                # ★★★ 結果を反映 ★★★
                                if alternative_found:
                                    pattern_type = alternative_found[0]

                                    if pattern_type in ['qty2_pair', 'qty3_4_pair', 'mixed']:
                                        # L/Rペアの場合
                                        _, lane1, pattern1, lane2, pattern2, target_parts = alternative_found

                                        # パターン1の全部品を追加
                                        for _, alt_row in df_parts_ref[
                                            (df_parts_ref['出庫先'] == lane1) &
                                            (df_parts_ref['パターン'].astype(str) == pattern1)
                                        ].iterrows():
                                            df_picking_list.loc[len(df_picking_list)] = {
                                                '部品番号': alt_row['部品番号'],
                                                '部品名称': alt_row['部品名称'],
                                                'PC': alt_row['pc'],
                                                '出庫先': lane1,
                                                'パターン': pattern1
                                            }

                                        # パターン2の全部品を追加
                                        for _, alt_row in df_parts_ref[
                                            (df_parts_ref['出庫先'] == lane2) &
                                            (df_parts_ref['パターン'].astype(str) == pattern2)
                                        ].iterrows():
                                            df_picking_list.loc[len(df_picking_list)] = {
                                                '部品番号': alt_row['部品番号'],
                                                '部品名称': alt_row['部品名称'],
                                                'PC': alt_row['pc'],
                                                '出庫先': lane2,
                                                'パターン': pattern2
                                            }

                                        # 720シートを更新
                                        self._update_720_sheet_pattern(ws720, lane1, pattern1)
                                        self._update_720_sheet_pattern(ws720, lane2, pattern2)

                                        success_count += 1
                                        print(
                                            f"            ✅ L/Rペア割り当て完了: {lane1}:{pattern1} + {lane2}:{pattern2}")

                                    elif pattern_type == 'qty1_single':
                                        # 単独パターンの場合
                                        _, alt_lane, alt_pattern, target_parts = alternative_found

                                        for _, alt_row in df_parts_ref[
                                            (df_parts_ref['出庫先'] == alt_lane) &
                                            (df_parts_ref['パターン'].astype(str) == alt_pattern)
                                        ].iterrows():
                                            df_picking_list.loc[len(df_picking_list)] = {
                                                '部品番号': alt_row['部品番号'],
                                                '部品名称': alt_row['部品名称'],
                                                'PC': alt_row['pc'],
                                                '出庫先': alt_lane,
                                                'パターン': alt_pattern
                                            }

                                        # 720シートを更新
                                        self._update_720_sheet_pattern(ws720, alt_lane, alt_pattern)

                                        success_count += 1
                                        print(f"            ✅ 単独パターン割り当て完了: {alt_lane}:{alt_pattern}")

                                    # target_laneを削除
                                    break
                                else:
                                    print(f"            ⚠️ 代替パターンが見つかりませんでした")

                    # L/Rペアの2行を残し、failed_partの他の行を削除
                    df_picking_list.drop(
                        df_picking_list[
                            (df_picking_list['部品番号'] == failed_part) &
                            (~df_picking_list['出庫先'].isin([lane1, lane2]))
                            ].index,
                        inplace=True
                    )

                    # 各レーンを1個ずつに調整
                    for lane in [lane1, lane2]:
                        lane_rows = df_picking_list[
                            (df_picking_list['部品番号'] == failed_part) &
                            (df_picking_list['出庫先'] == lane)
                            ]
                        if len(lane_rows) > 0:
                            # 最初の行を1個に設定
                            df_picking_list.loc[lane_rows.index[0], 'PC'] = 1
                            # 2行目以降を削除
                            if len(lane_rows) > 1:
                                df_picking_list.drop(lane_rows.index[1:], inplace=True)

                    success_count += 1
                    print(f"            ✅ L/Rペアで2個に調整完了")
                    continue

            # ===== ケース2: 余剰がある場合、孤立部品を探して代替パターンを検索 =====
            if len(lane_qty) > required_qty or sum(lane_qty.values()) > required_qty:
                print(f"            余剰検出 → 孤立部品の代替パターンを探索")

                # 各出庫先のパターン情報を取得
                lane_patterns = {}
                for _, row in current_assignments.iterrows():
                    lane = row['出庫先']
                    pattern = row['パターン']
                    if lane not in lane_patterns:
                        lane_patterns[lane] = pattern

                # 各出庫先の全部品を取得
                lane_all_parts = {}
                for lane, pattern in lane_patterns.items():
                    key = f"{lane}|{pattern}"
                    if key in pattern_dict:
                        lane_all_parts[lane] = set(pattern_dict[key].keys())

                # 孤立部品を検出（他の出庫先で使われていない部品）
                for target_lane in lane_patterns.keys():
                    if target_lane not in lane_all_parts:
                        continue

                    # この出庫先の全部品
                    target_parts = lane_all_parts[target_lane]

                    # 他の出庫先で使われている部品
                    other_parts = set()
                    for other_lane, parts in lane_all_parts.items():
                        if other_lane != target_lane:
                            other_parts.update(parts)

                    # 孤立部品（この出庫先にしかない部品）
                    isolated_parts = target_parts - other_parts - {failed_part}

                    if len(isolated_parts) == 0:
                        continue

                    print(f"            孤立部品検出 ({target_lane}): {isolated_parts}")

                    # 全出庫先から孤立部品の単独パターンを探索
                    alternative_found = None

                    # ★★★ 孤立部品を数量別にグループ化 ★★★
                    parts_qty_2 = set()  # 数量2個の部品
                    parts_qty_3_4 = set()  # 数量3-4個の部品
                    parts_qty_1 = set()  # 数量1個の部品

                    for isolated_part in isolated_parts:
                        # 必要数量を取得（構成表マトリックスから）
                        part_qty_rows = df_matrix[
                            (df_matrix.iloc[:, 6] == '_1') &
                            (df_matrix.iloc[:, 4].astype(str).str.replace('*', '',
                                                                          regex=False).str.strip() == isolated_part)
                            ]

                        if len(part_qty_rows) > 0:
                            required_qty = pd.to_numeric(part_qty_rows.iloc[0, 9], errors='coerce')
                            if pd.notna(required_qty):
                                required_qty = int(required_qty)

                                if required_qty == 2:
                                    parts_qty_2.add(isolated_part)
                                elif required_qty in [3, 4]:
                                    parts_qty_3_4.add(isolated_part)
                                else:
                                    parts_qty_1.add(isolated_part)
                            else:
                                parts_qty_1.add(isolated_part)
                        else:
                            parts_qty_1.add(isolated_part)

                    print(f"              数量2個: {len(parts_qty_2)}部品")
                    print(f"              数量3-4個: {len(parts_qty_3_4)}部品")
                    print(f"              数量1個: {len(parts_qty_1)}部品")

                    alternative_found = None

                    # ★★★ 優先度1: 数量2個 → 9L2A + 9R2A ペア ★★★
                    if len(parts_qty_2) > 0:
                        print(f"            優先探索: 数量2個部品を 9L2A + 9R2A ペアで探索...")

                        # 9L2Aのパターンを探す
                        pattern_9l2a = self._find_pattern_containing_parts_with_qty(
                            '9L2A', parts_qty_2, 1, df_parts_ref, df_picking_list
                        )

                        # 9R2Aのパターンを探す
                        pattern_9r2a = self._find_pattern_containing_parts_with_qty(
                            '9R2A', parts_qty_2, 1, df_parts_ref, df_picking_list
                        )

                        if pattern_9l2a and pattern_9r2a:
                            alternative_found = ('qty2_pair', '9L2A', pattern_9l2a, '9R2A', pattern_9r2a, parts_qty_2)
                            print(f"            ✅ 数量2個ペア発見: 9L2A:{pattern_9l2a} + 9R2A:{pattern_9r2a}")

                    # ★★★ 優先度2: 数量3-4個 → 7L2 + 7R2 ペア ★★★
                    if not alternative_found and len(parts_qty_3_4) > 0:
                        print(f"            次善策: 数量3-4個部品を 7L2 + 7R2 ペアで探索...")

                        # 7L2のパターンを探す
                        pattern_7l2 = self._find_pattern_containing_parts_any_qty(
                            '7L2', parts_qty_3_4, df_parts_ref, df_picking_list
                        )

                        # 7R2のパターンを探す
                        pattern_7r2 = self._find_pattern_containing_parts_any_qty(
                            '7R2', parts_qty_3_4, df_parts_ref, df_picking_list
                        )

                        if pattern_7l2 and pattern_7r2:
                            alternative_found = ('qty3_4_pair', '7L2', pattern_7l2, '7R2', pattern_7r2, parts_qty_3_4)
                            print(f"            ✅ 数量3-4個ペア発見: 7L2:{pattern_7l2} + 7R2:{pattern_7r2}")

                    # ★★★ 優先度3: 数量1個 → 単独パターン ★★★
                    if not alternative_found and len(parts_qty_1) > 0:
                        print(f"            次善策: 数量1個部品を単独パターンで探索...")

                        # 全出庫先から探索
                        for _, cand_row in df_parts_ref.iterrows():
                            alt_lane = cand_row['出庫先']
                            alt_pattern = str(cand_row['パターン']).strip()

                            pattern_rows = df_parts_ref[
                                (df_parts_ref['出庫先'] == alt_lane) &
                                (df_parts_ref['パターン'].astype(str) == alt_pattern)
                                ]

                            pattern_parts = set(pattern_rows['部品番号'].tolist())

                            # 数量1個の部品全てが含まれているかチェック
                            if parts_qty_1.issubset(pattern_parts):
                                # 未使用かチェック
                                already_used = df_picking_list[
                                    (df_picking_list['出庫先'] == alt_lane) &
                                    (df_picking_list['パターン'].astype(str) == alt_pattern)
                                    ]

                                if len(already_used) == 0:
                                    alternative_found = ('qty1_single', alt_lane, alt_pattern, parts_qty_1)
                                    print(f"            ✅ 数量1個パターン発見: {alt_lane}:{alt_pattern}")
                                    break

                    # ★★★ 混合ケース: 数量2個 + 数量1個 ★★★
                    if not alternative_found and len(parts_qty_2) > 0 and len(parts_qty_1) > 0:
                        print(f"            次善策: 数量2個+1個を混合パターンで探索...")

                        # 9L2A側: 数量2個(各1個) + 数量1個の一部
                        # 9R2A側: 数量2個(各1個) + 数量1個の残り

                        # すべての数量1個部品の分割パターンを試す
                        from itertools import combinations
                        parts_1_list = list(parts_qty_1)

                        for split_count in range(len(parts_1_list) + 1):
                            for group1_parts in combinations(parts_1_list, split_count):
                                group1 = parts_qty_2 | set(group1_parts)
                                group2 = parts_qty_2 | (parts_qty_1 - set(group1_parts))

                                # 9L2Aで group1 を探す（数量2個は1個ずつ）
                                pattern_9l2a = self._find_mixed_pattern(
                                    '9L2A', parts_qty_2, 1, set(group1_parts), 1, df_parts_ref, df_picking_list
                                )

                                # 9R2Aで group2 を探す（数量2個は1個ずつ）
                                pattern_9r2a = self._find_mixed_pattern(
                                    '9R2A', parts_qty_2, 1, parts_qty_1 - set(group1_parts), 1, df_parts_ref,
                                    df_picking_list
                                )

                                if pattern_9l2a and pattern_9r2a:
                                    alternative_found = ('mixed', '9L2A', pattern_9l2a, '9R2A', pattern_9r2a,
                                                         parts_qty_2 | parts_qty_1)
                                    print(f"            ✅ 混合パターン発見: 9L2A:{pattern_9l2a} + 9R2A:{pattern_9r2a}")
                                    break

                            if alternative_found:
                                break

                    # ★★★ 結果を反映 ★★★
                    if alternative_found:
                        pattern_type = alternative_found[0]

                        if pattern_type in ['qty2_pair', 'qty3_4_pair', 'mixed']:
                            # L/Rペアの場合
                            _, lane1, pattern1, lane2, pattern2, target_parts = alternative_found

                            # パターン1の全部品を追加
                            for _, alt_row in df_parts_ref[
                                (df_parts_ref['出庫先'] == lane1) &
                                (df_parts_ref['パターン'].astype(str) == pattern1)
                            ].iterrows():
                                df_picking_list.loc[len(df_picking_list)] = {
                                    '部品番号': alt_row['部品番号'],
                                    '部品名称': alt_row['部品名称'],
                                    'PC': alt_row['pc'],
                                    '出庫先': lane1,
                                    'パターン': pattern1
                                }

                            # パターン2の全部品を追加
                            for _, alt_row in df_parts_ref[
                                (df_parts_ref['出庫先'] == lane2) &
                                (df_parts_ref['パターン'].astype(str) == pattern2)
                            ].iterrows():
                                df_picking_list.loc[len(df_picking_list)] = {
                                    '部品番号': alt_row['部品番号'],
                                    '部品名称': alt_row['部品名称'],
                                    'PC': alt_row['pc'],
                                    '出庫先': lane2,
                                    'パターン': pattern2
                                }

                            # 720シートを更新
                            self._update_720_sheet_pattern(ws720, lane1, pattern1)
                            self._update_720_sheet_pattern(ws720, lane2, pattern2)

                            success_count += 1
                            print(f"            ✅ L/Rペア割り当て完了: {lane1}:{pattern1} + {lane2}:{pattern2}")

                        elif pattern_type == 'qty1_single':
                            # 単独パターンの場合
                            _, alt_lane, alt_pattern, target_parts = alternative_found

                            for _, alt_row in df_parts_ref[
                                (df_parts_ref['出庫先'] == alt_lane) &
                                (df_parts_ref['パターン'].astype(str) == alt_pattern)
                            ].iterrows():
                                df_picking_list.loc[len(df_picking_list)] = {
                                    '部品番号': alt_row['部品番号'],
                                    '部品名称': alt_row['部品名称'],
                                    'PC': alt_row['pc'],
                                    '出庫先': alt_lane,
                                    'パターン': alt_pattern
                                }

                            # 720シートを更新
                            self._update_720_sheet_pattern(ws720, alt_lane, alt_pattern)

                            success_count += 1
                            print(f"            ✅ 単独パターン割り当て完了: {alt_lane}:{alt_pattern}")

                        # target_laneを削除
                        break
                    else:
                        print(f"            ⚠️ 代替パターンが見つかりませんでした")

        return success_count > 0

    def _find_optimal_pattern_combination(self, failed_parts, df_verification, df_parts_ref, df_picking_list,
                                          df_matrix):
        """
        全ての不一致部品に対して、最適なパターン組み合わせを見つける（大改修版）

        Args:
            failed_parts: 不一致部品のリスト
            df_verification: 検証用DataFrame
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト
            df_matrix: 構成表マトリックス

        Returns:
            最適なパターン組み合わせの辞書 {部品番号: [(出庫先, パターン, 数量), ...]}
        """
        print(f"\n      🔍 最適パターン組み合わせ探索（大改修版）")

        # ===== ステップ1: 各部品の必要数量を取得 =====
        required_quantities = {}
        for part in failed_parts:
            qty = df_verification[df_verification['部品番号'] == part]['PC'].iloc[0]
            required_quantities[part] = qty
            print(f"         {part}: {qty}個必要")

        # ===== ステップ2: 各部品の候補パターンを収集 =====
        print(f"\n      📋 候補パターン収集中...")
        candidates = {}

        for part in failed_parts:
            part_candidates = []

            # 参照DBから全候補を取得
            part_patterns = df_parts_ref[df_parts_ref['部品番号'] == part]

            for _, row in part_patterns.iterrows():
                lane = row['出庫先']
                pattern = str(row['パターン']).strip()
                pc = int(row['pc']) if pd.notna(row['pc']) else 1

                # 未使用かチェック
                already_used = df_picking_list[
                    (df_picking_list['出庫先'] == lane) &
                    (df_picking_list['パターン'].astype(str) == pattern)
                    ]

                if len(already_used) == 0:
                    part_candidates.append({
                        'lane': lane,
                        'pattern': pattern,
                        'pc': pc
                    })

            candidates[part] = part_candidates
            print(f"         {part}: {len(part_candidates)}個の候補")

        # ===== ステップ3: 最適な組み合わせを探索 =====
        print(f"\n      🎯 最適組み合わせ探索中...")

        best_combination = self._search_best_combination(
            candidates, required_quantities, df_parts_ref
        )

        if best_combination:
            print(f"      ✅ 最適組み合わせ発見！")
            for part, patterns in best_combination.items():
                total_qty = sum(p['pc'] for p in patterns)
                print(f"         {part}: {total_qty}個")
                for p in patterns:
                    print(f"           - {p['lane']}:{p['pattern']} ({p['pc']}個)")
        else:
            print(f"      ❌ 最適組み合わせが見つかりませんでした")

        return best_combination

    def _search_best_combination(self, candidates, required_quantities, df_parts_ref):
        """
        最適なパターン組み合わせを探索

        優先順位:
        1. 数量2個の部品 → L/Rペアを優先
        2. 必要数量との完全一致
        3. パターン数の少なさ
        """
        from itertools import combinations, product

        # 部品を数量別にグループ化
        parts_qty_2 = {p: q for p, q in required_quantities.items() if q == 2}
        parts_qty_other = {p: q for p, q in required_quantities.items() if q != 2}

        best_combination = {}

        # ===== 数量2個の部品を優先処理（L/Rペア） =====
        for part, required_qty in parts_qty_2.items():
            part_candidates = candidates[part]

            # L/Rペアを探す
            lr_pair_found = None

            for i, cand1 in enumerate(part_candidates):
                for cand2 in part_candidates[i + 1:]:
                    if self._is_lr_pair(cand1['lane'], cand2['lane']):
                        if cand1['pc'] == 1 and cand2['pc'] == 1:
                            lr_pair_found = [cand1, cand2]
                            break
                if lr_pair_found:
                    break

            if lr_pair_found:
                best_combination[part] = lr_pair_found
                print(f"         ✅ {part}: L/Rペア発見 {lr_pair_found[0]['lane']} + {lr_pair_found[1]['lane']}")
            else:
                # L/Rペアがない場合、数量一致を探す
                for cand in part_candidates:
                    if cand['pc'] == required_qty:
                        best_combination[part] = [cand]
                        break

        # ===== 数量2個以外の部品を処理 =====
        for part, required_qty in parts_qty_other.items():
            part_candidates = candidates[part]

            # 数量完全一致を優先
            for cand in part_candidates:
                if cand['pc'] == required_qty:
                    best_combination[part] = [cand]
                    break

            # 見つからない場合、複数組み合わせを試す
            if part not in best_combination:
                for r in range(1, min(4, len(part_candidates) + 1)):
                    for combo in combinations(part_candidates, r):
                        if sum(c['pc'] for c in combo) == required_qty:
                            best_combination[part] = list(combo)
                            break
                    if part in best_combination:
                        break

        # 全ての部品に解が見つかったかチェック
        if len(best_combination) == len(required_quantities):
            return best_combination
        else:
            return None













    def _find_pattern_containing_parts_with_qty(self, lane, required_parts, required_qty_per_part, df_parts_ref,
                                                df_picking_list):
        """
        指定された部品を全て含み、各部品の数量が一致するパターンを探す

        Args:
            lane: 出庫先
            required_parts: 必要な部品番号のセット
            required_qty_per_part: 各部品の必要数量
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト

        Returns:
            パターン番号（見つからない場合はNone）
        """
        # この出庫先の全パターンを取得
        lane_patterns = df_parts_ref[df_parts_ref['出庫先'] == lane]['パターン'].unique()

        for pattern in lane_patterns:
            pattern = str(pattern).strip()

            # このパターンの全部品を取得
            pattern_rows = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ]

            # 必要な部品が全て含まれているかチェック
            pattern_parts = {}
            for _, row in pattern_rows.iterrows():
                part = str(row['部品番号']).strip()
                qty = int(row['pc']) if pd.notna(row['pc']) else 0
                pattern_parts[part] = qty

            # 全ての必要部品が存在し、数量が一致するかチェック
            all_match = True
            for part in required_parts:
                if part not in pattern_parts or pattern_parts[part] != required_qty_per_part:
                    all_match = False
                    break

            if not all_match:
                continue

            # 未使用かチェック
            already_used = df_picking_list[
                (df_picking_list['出庫先'] == lane) &
                (df_picking_list['パターン'].astype(str) == pattern)
                ]

            if len(already_used) == 0:
                return pattern

        return None

    def _find_pattern_containing_parts_any_qty(self, lane, required_parts, df_parts_ref, df_picking_list):
        """
        指定された部品を全て含むパターンを探す（数量は問わない）

        Args:
            lane: 出庫先
            required_parts: 必要な部品番号のセット
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト

        Returns:
            パターン番号（見つからない場合はNone）
        """
        # この出庫先の全パターンを取得
        lane_patterns = df_parts_ref[df_parts_ref['出庫先'] == lane]['パターン'].unique()

        for pattern in lane_patterns:
            pattern = str(pattern).strip()

            # このパターンの全部品を取得
            pattern_rows = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ]

            pattern_parts = set(pattern_rows['部品番号'].tolist())

            # 必要な部品が全て含まれているかチェック
            if required_parts.issubset(pattern_parts):
                # 未使用かチェック
                already_used = df_picking_list[
                    (df_picking_list['出庫先'] == lane) &
                    (df_picking_list['パターン'].astype(str) == pattern)
                    ]

                if len(already_used) == 0:
                    return pattern

        return None

    def _find_mixed_pattern(self, lane, parts_qty_2, qty_2, parts_qty_1, qty_1, df_parts_ref, df_picking_list):
        """
        数量2個の部品と数量1個の部品が混在するパターンを探す

        Args:
            lane: 出庫先
            parts_qty_2: 数量2個の部品セット
            qty_2: parts_qty_2の各部品の必要数量
            parts_qty_1: 数量1個の部品セット
            qty_1: parts_qty_1の各部品の必要数量
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト

        Returns:
            パターン番号（見つからない場合はNone）
        """
        # この出庫先の全パターンを取得
        lane_patterns = df_parts_ref[df_parts_ref['出庫先'] == lane]['パターン'].unique()

        for pattern in lane_patterns:
            pattern = str(pattern).strip()

            # このパターンの全部品を取得
            pattern_rows = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ]

            # 部品ごとの数量を取得
            pattern_parts = {}
            for _, row in pattern_rows.iterrows():
                part = str(row['部品番号']).strip()
                qty = int(row['pc']) if pd.notna(row['pc']) else 0
                pattern_parts[part] = qty

            # 全ての条件を満たすかチェック
            all_match = True

            # 数量2個の部品チェック
            for part in parts_qty_2:
                if part not in pattern_parts or pattern_parts[part] != qty_2:
                    all_match = False
                    break

            # 数量1個の部品チェック
            if all_match:
                for part in parts_qty_1:
                    if part not in pattern_parts or pattern_parts[part] != qty_1:
                        all_match = False
                        break

            if not all_match:
                continue

            # 未使用かチェック
            already_used = df_picking_list[
                (df_picking_list['出庫先'] == lane) &
                (df_picking_list['パターン'].astype(str) == pattern)
                ]

            if len(already_used) == 0:
                return pattern

        return None

    def _find_pattern_containing_parts_with_qty(self, lane, required_parts, required_qty_per_part, df_parts_ref,
                                                df_picking_list):
        """
        指定された部品を全て含み、各部品の数量が一致するパターンを探す

        Args:
            lane: 出庫先
            required_parts: 必要な部品番号のセット
            required_qty_per_part: 各部品の必要数量
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト

        Returns:
            パターン番号（見つからない場合はNone）
        """
        # この出庫先の全パターンを取得
        lane_patterns = df_parts_ref[df_parts_ref['出庫先'] == lane]['パターン'].unique()

        for pattern in lane_patterns:
            pattern = str(pattern).strip()

            # このパターンの全部品を取得
            pattern_rows = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ]

            # 必要な部品が全て含まれているかチェック
            pattern_parts = {}
            for _, row in pattern_rows.iterrows():
                part = str(row['部品番号']).strip()
                qty = int(row['pc']) if pd.notna(row['pc']) else 0
                pattern_parts[part] = qty

            # 全ての必要部品が存在し、数量が一致するかチェック
            all_match = True
            for part in required_parts:
                if part not in pattern_parts or pattern_parts[part] != required_qty_per_part:
                    all_match = False
                    break

            if not all_match:
                continue

            # 未使用かチェック
            already_used = df_picking_list[
                (df_picking_list['出庫先'] == lane) &
                (df_picking_list['パターン'].astype(str) == pattern)
                ]

            if len(already_used) == 0:
                return pattern

        return None

    def _find_pattern_containing_parts_any_qty(self, lane, required_parts, df_parts_ref, df_picking_list):
        """
        指定された部品を全て含むパターンを探す（数量は問わない）

        Args:
            lane: 出庫先
            required_parts: 必要な部品番号のセット
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト

        Returns:
            パターン番号（見つからない場合はNone）
        """
        # この出庫先の全パターンを取得
        lane_patterns = df_parts_ref[df_parts_ref['出庫先'] == lane]['パターン'].unique()

        for pattern in lane_patterns:
            pattern = str(pattern).strip()

            # このパターンの全部品を取得
            pattern_rows = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ]

            pattern_parts = set(pattern_rows['部品番号'].tolist())

            # 必要な部品が全て含まれているかチェック
            if required_parts.issubset(pattern_parts):
                # 未使用かチェック
                already_used = df_picking_list[
                    (df_picking_list['出庫先'] == lane) &
                    (df_picking_list['パターン'].astype(str) == pattern)
                    ]

                if len(already_used) == 0:
                    return pattern

        return None

    def _find_mixed_pattern(self, lane, parts_qty_2, qty_2, parts_qty_1, qty_1, df_parts_ref, df_picking_list):
        """
        数量2個の部品と数量1個の部品が混在するパターンを探す

        Args:
            lane: 出庫先
            parts_qty_2: 数量2個の部品セット
            qty_2: parts_qty_2の各部品の必要数量
            parts_qty_1: 数量1個の部品セット
            qty_1: parts_qty_1の各部品の必要数量
            df_parts_ref: A部品ピッキング参照DB
            df_picking_list: 現在のピッキングリスト

        Returns:
            パターン番号（見つからない場合はNone）
        """
        # この出庫先の全パターンを取得
        lane_patterns = df_parts_ref[df_parts_ref['出庫先'] == lane]['パターン'].unique()

        for pattern in lane_patterns:
            pattern = str(pattern).strip()

            # このパターンの全部品を取得
            pattern_rows = df_parts_ref[
                (df_parts_ref['出庫先'] == lane) &
                (df_parts_ref['パターン'].astype(str) == pattern)
                ]

            # 部品ごとの数量を取得
            pattern_parts = {}
            for _, row in pattern_rows.iterrows():
                part = str(row['部品番号']).strip()
                qty = int(row['pc']) if pd.notna(row['pc']) else 0
                pattern_parts[part] = qty

            # 全ての条件を満たすかチェック
            all_match = True

            # 数量2個の部品チェック
            for part in parts_qty_2:
                if part not in pattern_parts or pattern_parts[part] != qty_2:
                    all_match = False
                    break

            # 数量1個の部品チェック
            if all_match:
                for part in parts_qty_1:
                    if part not in pattern_parts or pattern_parts[part] != qty_1:
                        all_match = False
                        break

            if not all_match:
                continue

            # 未使用かチェック
            already_used = df_picking_list[
                (df_picking_list['出庫先'] == lane) &
                (df_picking_list['パターン'].astype(str) == pattern)
                ]

            if len(already_used) == 0:
                return pattern

        return None

    def _update_720_sheet_pattern(self, ws720, lane, pattern):
        """720システム入力シートのパターンを更新"""

        # レーン名を検索（B21:K21, B25:K25）
        target_cell = None

        for row in [21, 25]:
            for col in range(2, 12):
                cell_value = str(ws720.cell(row, col).value or "").strip()
                if cell_value == lane:
                    target_cell = (row, col)
                    break
            if target_cell:
                break

        if target_cell:
            row, col = target_cell
            # パターンを2桁表示に変換
            if pattern.isdigit():
                pattern = pattern.zfill(2)
            ws720.cell(row + 1, col, pattern)
            ws720.cell(row + 1, col).number_format = '@'
            print(f"              720シート更新: {lane} → パターン{pattern}")

    # ============================================================================
    # メイン実行
    # ============================================================================
    def run(self, is_a_line=True):
        """メイン実行"""
        print("\n" + "=" * 80)
        print("流用元無しピッキングリスト自動生成プログラム")
        print("=" * 80)
        print(f"実行日時: {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}")
        print(f"処理モード: {'Aライン（CM + A部品）' if is_a_line else 'Cライン（CMのみ）'}")
        print("=" * 80)

        try:
            target_file = self.find_target_file()
            self.validate_file_structure(target_file)
            df_matrix, frame_number, original_file = self.load_matrix(target_file)
            df_cm_ref = self.load_cm_master()
            df_parts_ref = self.load_parts_master() if is_a_line else pd.DataFrame()
            df_cm = self.create_cm_picking(df_matrix, df_cm_ref)
            df_a_picking = self.create_a_picking_list(df_matrix, df_parts_ref) if is_a_line else pd.DataFrame()

            output_file = self.save_to_excel(
                df_matrix, df_cm_ref, df_parts_ref, df_cm, df_a_picking,
                frame_number, original_file, is_a_line
            )

            print("\n" + "=" * 80)
            print("✅✅✅ 処理完了！ ✅✅✅")
            print("=" * 80)
            print(f"📄 出力ファイル: {output_file.name}")
            print(f"📁 保存先: {output_file.parent}")
            print(f"📊 CMピッキング: {len(df_cm)}部品")
            if is_a_line:
                print(f"📊 A部品ピッキング: {len(df_a_picking)}部品")
            print("=" * 80)

        except Exception as e:
            print("\n" + "=" * 80)
            print("❌ エラー発生")
            print("=" * 80)
            print(f"エラー内容: {e}")
            import traceback
            traceback.print_exc()

    # ============================================================================
    # 除外表による行削除
    # ============================================================================
    def _delete_rows_by_exclusion_list(self, df_cm_ref):
        """
        除外表に基づいてCMピッキング参照DBの行を削除

        【説明】
        外部ファイル「流用元無しピック_除外表.xlsx」のA,B,C列と、
        CMピッキング参照DBのレーン番号・出庫先レーン・出庫先パターンを照合し、
        完全一致する行を削除します。

        【除外表ファイル】
        パス: C:\\temp\\Newピッキング_対象照会\\参照先\\流用元無しピック_除外表.xlsx
        シート: 1番目のシート
        列構成: A列=レーン番号, B列=出庫先レーン, C列=出庫先パターン

        Args:
            df_cm_ref (DataFrame): CMピッキング参照DB

        Returns:
            DataFrame: 除外処理後のCMピッキング参照DB
        """
        print("\n  🔍 除外表による行削除処理...")

        exclusion_file = self.master_dir / "流用元無しピック_除外表.xlsx"

        if not exclusion_file.exists():
            print(f"      ℹ️ 除外表ファイルが見つかりません（スキップ）")
            print(f"         {exclusion_file}")
            return df_cm_ref

        try:
            df_exclusion = pd.read_excel(exclusion_file, dtype=str).fillna("")

            if len(df_exclusion.columns) < 3:
                print(f"      ⚠️ 除外表の列数が不足しています（列数: {len(df_exclusion.columns)}）")
                return df_cm_ref

            df_exclusion = df_exclusion.iloc[:, :3].copy()
            df_exclusion.columns = ['レーン番号', '出庫先レーン', '出庫先パターン']

            for col in df_exclusion.columns:
                df_exclusion[col] = df_exclusion[col].apply(self.clean_value)

            df_exclusion = df_exclusion[
                (df_exclusion['レーン番号'] != "") &
                (df_exclusion['出庫先レーン'] != "") &
                (df_exclusion['出庫先パターン'] != "")
                ].copy()

            df_exclusion = df_exclusion[
                df_exclusion['レーン番号'].astype(str).str.strip() != "レーン番号"
                ].copy()

            if len(df_exclusion) == 0:
                print(f"      ℹ️ 除外対象データなし")
                return df_cm_ref

            print(f"      📋 除外表: {len(df_exclusion)}件読込")

            exclusion_set = set()
            for _, row in df_exclusion.iterrows():
                key = (
                    str(row['レーン番号']).strip(),
                    str(row['出庫先レーン']).strip(),
                    str(row['出庫先パターン']).strip()
                )
                exclusion_set.add(key)

            before_count = len(df_cm_ref)

            df_cm_ref['_exclude_key'] = df_cm_ref.apply(
                lambda row: (
                    str(row['レーン番号']).strip(),
                    str(row['出庫先レーン']).strip(),
                    str(row['出庫先パターン']).strip()
                ),
                axis=1
            )

            df_cm_ref = df_cm_ref[~df_cm_ref['_exclude_key'].isin(exclusion_set)].copy()
            df_cm_ref = df_cm_ref.drop(columns=['_exclude_key'])

            after_count = len(df_cm_ref)
            deleted_count = before_count - after_count

            if deleted_count > 0:
                print(f"      ✅ {deleted_count}行を削除しました（残り: {after_count}行）")
                self.logger.add_step(
                    "除外表による行削除",
                    f"{deleted_count}行削除、残り{after_count}行"
                )
            else:
                print(f"      ℹ️ 削除対象なし（全{before_count}行を保持）")

            return df_cm_ref

        except Exception as e:
            print(f"      ⚠️ 除外表処理エラー: {e}")
            self.logger.add_warning("除外表処理エラー", str(e))
            import traceback
            traceback.print_exc()
            return df_cm_ref

# ================================================================================
# メイン実行
# ================================================================================
if __name__ == "__main__":
    print("\n" + "=" * 80)
    print("処理ライン選択")
    print("=" * 80)
    print("1: Aライン（CMピッキング + A部品ピッキング）")
    print("2: Cライン（CMピッキングのみ）")

    choice = input("\n選択してください (1/2): ").strip()

    generator = PickingListGenerator()

    if choice == "1":
        print("\n>>> Aライン処理を開始します")
        generator.run(is_a_line=True)
    elif choice == "2":
        print("\n>>> Cライン処理を開始します")
        generator.run(is_a_line=False)
    else:
        print("❌ 無効な選択です")

    input("\nEnterキーを押して終了...")